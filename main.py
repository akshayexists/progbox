from runsim import *
from exportcleaner import *

#look at all available teams
teams = []    #change to list of teams if specific teams are to be tested eg. ['GSW', 'BOS', 'NOL']

seed = 69
print("Master seed:", seed)
runs = 1000
path = 'outputs/raw/'
data = exportcleaner(export_file='data/export.json', teaminfo_file='data/teaminfo.json', teams = teams)
data.to_csv(path+'inputs.csv')
#print(data)

sim = runsim(seed=seed)
df = sim.PROGEMUP(data, runs=runs)
df.to_csv(path+'outputs.csv')

print(f'Written to {path}')

#---------------------- exporting to excel for Further Analysis ----------------------
# modify this if you want to change how aggregation is done or to add more analysis types
# I could've folded this into a method of the runsim class itself, but I find it better to keep
# analysis separate from data collection.

import pandas as pd
import json
from openpyxl import load_workbook
from openpyxl.worksheet.table import TableColumn
from openpyxl.utils import get_column_letter, range_boundaries

dataraw = pd.read_csv(path+'outputs.csv', index_col=0)
data = dataraw.drop(['Team', 'RunSeed', 'Run', 'AboveBaseline', 'Value', 'PlayerID'], axis=1).sort_values(by='Name')
meaned = (
    data.groupby("Name", sort=True)
        .agg(
            Ovr=("Ovr", "mean"),
            Ovr_min=("Ovr", "min"),
            Ovr_max=("Ovr", "max"),
            Ovr_q10=("Ovr", lambda s: s.quantile(0.10)),
            Ovr_q25=("Ovr", lambda s: s.quantile(0.25)),
            Ovr_q75=("Ovr", lambda s: s.quantile(0.75)),
            Ovr_q90=("Ovr", lambda s: s.quantile(0.90)),
            **{f"{col}": (col, "mean") 
               for col in data.columns if col not in {"Name", "Ovr"}}
        )
        .reset_index()
)

# --- Load and process godprogs.json ---
godprogs = (
    pd.read_json(path+'godprogs.json')
      .drop(['RunSeed', 'OVR', 'Age'], axis=1)
      .groupby('Name', sort=True)
      .mean()
      .reset_index()
)
godprogs.columns = ['Name', 'GodProg Average', 'GodProg Chance']

# --- Load and process superlucky.json ---
with open(path+'superlucky.json', 'rb') as f:
    superlucky = json.load(f)

superlucky = (
    pd.DataFrame(list(superlucky.items()), columns=['Name', 'GodProgCount'])
    .sort_values('Name')
    .reset_index(drop=True)
)

# --- Merge GodProgCount into godprogs ---
godprogs = godprogs.merge(superlucky, on='Name', how='left')

# Optional: fill missing GodProgCount with 0
godprogs['GodProgCount'] = godprogs['GodProgCount'].fillna(0)

teams_unique = dataraw[["Name", "Team"]].drop_duplicates(subset="Name")

aggregated = meaned.merge(
    godprogs,
    on="Name",
    how="left"   # keep all rows from big df
)

# Merge safely
aggregated = aggregated.merge(
    teams_unique,
    on="Name",
    how="left"
)

aggregated = aggregated[['Name', 'Team', 'Age', 'Baseline', 'Ovr', 'Delta', 'PctChange', 'Ovr_min', 'Ovr_max', 'Ovr_q10', 'Ovr_q25', 'Ovr_q75', 'Ovr_q90', 'PER', 'GodProg Average', 'GodProg Chance', 'GodProgCount', 
                        'dIQ', 'Dnk', 'Drb', 'End', '2Pt', 'FT', 'Ins', 'Jmp', 'oIQ', 'Pss', 'Reb', 'Spd', 'Str', '3Pt', 'Hgt']]


# aggregated is your pandas DataFrame
wb = load_workbook("outputs/analysis.xlsx")
ws = wb["aggregated"]

# target size (including header row)
nrows = len(aggregated) + 1
ncols = len(aggregated.columns)
end_col_letter = get_column_letter(ncols)
new_ref = f"A1:{end_col_letter}{nrows}"

# --- 1) Find candidate table(s) to update ---
candidate_tables = []
# ws._tables is a dict of name -> Table objects
for t in ws._tables.values():
    min_col, min_row, max_col, max_row = range_boundaries(t.ref)
    # pick tables that start at the header row (min_row == 1) as likely target
    if min_row == 1:
        candidate_tables.append(t)

# fallback: if no table with min_row==1 found, pick the first table if any
if not candidate_tables and ws._tables:
    candidate_tables = list(ws._tables.values())[:1]

# --- 2) Update table ref and columns for each candidate ---
for t in candidate_tables:
    # set the new ref (range) for the table
    t.ref = new_ref
    # rebuild tableColumns to match DataFrame column names (keeps the table structure valid)
    t.tableColumns = [TableColumn(id=i+1, name=str(col)) for i, col in enumerate(aggregated.columns)]

# --- 3) Clear only the relevant cell values (don't delete rows/cols!) ---
# Clear a bounding box covering the old used area and new area to be safe
max_row = max(ws.max_row, nrows)
max_col = max(ws.max_column, ncols)
for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
    for cell in row:
        cell.value = None

# --- 4) Write headers and data ---
for c_idx, col_name in enumerate(aggregated.columns, start=1):
    ws.cell(row=1, column=c_idx, value=col_name)

for r_idx, row in enumerate(aggregated.values, start=2):
    for c_idx, value in enumerate(row, start=1):
        ws.cell(row=r_idx, column=c_idx, value=value)

wb.save("outputs/analysis.xlsx")
