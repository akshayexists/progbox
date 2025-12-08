import os
import json
import shutil
import datetime
import traceback
import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.table import TableColumn
from openpyxl.utils import get_column_letter, range_boundaries
from runsim import runsim
from exportcleaner import exportcleaner
from visualize import create_plots


def generate_excel_analysis(output_dir, outputs_csv_path, raw_dir, template_path="outputs/analysis.xlsx"):
    """
    Generate analysis.xlsx in the output directory by aggregating simulation results.
    
    Copies the template Excel file and populates it with aggregated player statistics,
    God Progression data, and percentile distributions.
    """
    # Copy template to output directory
    dest_path = f"{output_dir}/analysis.xlsx"
    shutil.copy(template_path, dest_path)
    
    # Load simulation results
    dataraw = pd.read_csv(outputs_csv_path, index_col=0)
    data = dataraw.drop(['Team', 'RunSeed', 'Run', 'AboveBaseline', 'Value', 'PlayerID'], axis=1).sort_values(by='Name')
    
    # Aggregate statistics per player
    meaned = (
        data.groupby("Name", sort=True)
            .agg(
                Ovr=("Ovr", "mean"),
                Ovr_min=("Ovr", "min"),
                Ovr_max=("Ovr", "max"),
                Ovr_q10=("Ovr", lambda s: s.quantile(0.10)),
                Ovr_q25=("Ovr", lambda s: s.quantile(0.25)),
                Ovr_q75=("Ovr", lambda s: s.quantile(0.75)),
                Ovr_q90=("Ovr", lambda s: s.quantile(0.90)),
                **{f"{col}": (col, "mean") 
                   for col in data.columns if col not in {"Name", "Ovr"}}
            )
            .reset_index()
    )
    
    # Load and process godprogs.json
    try:
        godprogs = pd.read_json(f"{raw_dir}/godprogs.json")
        
        if godprogs.empty:
            godprogs = pd.DataFrame(columns=['Name', 'GodProg Average', 'GodProg Chance'])
        else:
            required_cols = ['RunSeed', 'OVR', 'Age', 'Name']
            missing_cols = [col for col in required_cols if col not in godprogs.columns]
            if missing_cols:
                raise ValueError(f"godprogs.json is missing required columns: {missing_cols}")
            
            godprogs = (
                godprogs.drop(['RunSeed', 'OVR', 'Age'], axis=1)
                       .groupby('Name', sort=True)
                       .mean()
                       .reset_index()
            )
            godprogs.columns = ['Name', 'GodProg Average', 'GodProg Chance']
            
    except FileNotFoundError:
        print(f"Warning: {raw_dir}/godprogs.json not found. Using empty GodProg data.")
        godprogs = pd.DataFrame(columns=['Name', 'GodProg Average', 'GodProg Chance'])
    except (ValueError, json.JSONDecodeError) as e:
        print(f"Warning: Failed to parse {raw_dir}/godprogs.json: {e}. Using empty GodProg data.")
        godprogs = pd.DataFrame(columns=['Name', 'GodProg Average', 'GodProg Chance'])
    
    # Load and process superlucky.json
    with open(f"{raw_dir}/superlucky.json", 'r', encoding='utf-8') as f:
        superlucky = json.load(f)
    
    superlucky = (
        pd.DataFrame(list(superlucky.items()), columns=['Name', 'GodProgCount'])
        .sort_values('Name')
        .reset_index(drop=True)
    )
    
    # Merge GodProgCount into godprogs
    godprogs = godprogs.merge(superlucky, on='Name', how='left')
    # Players in godprogs have at least 1 god progression; superlucky tracks multiple (2+)
    godprogs['GodProgCount'] = godprogs['GodProgCount'].fillna(1)
    
    teams_unique = dataraw[["Name", "Team"]].drop_duplicates(subset="Name")
    
    aggregated = meaned.merge(godprogs, on="Name", how="left")
    aggregated = aggregated.merge(teams_unique, on="Name", how="left")
    
    aggregated = aggregated[['Name', 'Team', 'Age', 'Baseline', 'Ovr', 'Delta', 'PctChange', 
                             'Ovr_min', 'Ovr_max', 'Ovr_q10', 'Ovr_q25', 'Ovr_q75', 'Ovr_q90', 
                             'PER', 'GodProg Average', 'GodProg Chance', 'GodProgCount', 
                             'dIQ', 'Dnk', 'Drb', 'End', '2Pt', 'FT', 'Ins', 'Jmp', 'oIQ', 
                             'Pss', 'Reb', 'Spd', 'Str', '3Pt', 'Hgt']]
    
    # Update Excel file
    wb = load_workbook(dest_path)
    ws = wb["aggregated"]
    
    nrows = len(aggregated) + 1
    ncols = len(aggregated.columns)
    end_col_letter = get_column_letter(ncols)
    new_ref = f"A1:{end_col_letter}{nrows}"
    
    # Find and update table(s)
    candidate_tables = []
    for t in ws._tables.values():
        min_col, min_row, max_col, max_row = range_boundaries(t.ref)
        if min_row == 1:
            candidate_tables.append(t)
    
    if not candidate_tables and ws._tables:
        candidate_tables = list(ws._tables.values())[:1]
    
    for t in candidate_tables:
        t.ref = new_ref
        t.tableColumns = [TableColumn(id=i+1, name=str(col)) for i, col in enumerate(aggregated.columns)]
    
    # Clear existing values
    clear_max_row = max(ws.max_row, nrows)
    clear_max_col = max(ws.max_column, ncols)
    for row in ws.iter_rows(min_row=1, max_row=clear_max_row, min_col=1, max_col=clear_max_col):
        for cell in row:
            cell.value = None
    
    # Write headers and data
    for c_idx, col_name in enumerate(aggregated.columns, start=1):
        ws.cell(row=1, column=c_idx, value=col_name)
    
    for r_idx, row in enumerate(aggregated.values, start=2):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)
    
    wb.save(dest_path)
    print(f"Excel analysis saved to {dest_path}")

def run_benchmark(export_file='data/export.json', teaminfo_file='data/teaminfo.json', teams=None, runs=1000, seed=69):
    """
    Execute a complete benchmark run with organized output.
    
    Creates a timestamped folder containing:
    - metadata.json: League info from export
    - outputs.csv: Full simulation results
    - analysis.xlsx: Aggregated Excel report
    - raw/: inputs.csv, godprogs.json, superlucky.json
    - plots/: Visualization PNGs
    """
    # Setup output directory
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    output_dir = f"outputs/run_{timestamp}"
    raw_dir = f"{output_dir}/raw"
    try:
        os.makedirs(output_dir, exist_ok=True)
        os.makedirs(raw_dir, exist_ok=True)
    except OSError as e:
        raise RuntimeError(f"Failed to create output directories: {e}") from e    
    print(f"Starting benchmark run. Output directory: {output_dir}")
    print(f"Master seed: {seed}")
    
    # Load data and extract metadata
    data, metadata = exportcleaner(export_file=export_file, teaminfo_file=teaminfo_file, teams=teams or [])
    
    # Save inputs and metadata
    data.to_csv(f"{raw_dir}/inputs.csv")
    # Add simulation parameters to metadata
    metadata['master_seed'] = seed
    metadata['runs'] = runs
    metadata['timestamp'] = timestamp
    with open(f"{output_dir}/metadata.json", "w") as f:
        json.dump(metadata, f, indent=4)
        
    print(f"Loaded {len(data)} players from {metadata.get('league_name', 'Unknown League')}")
    
    # Run simulation
    sim = runsim(seed=seed)
    results_df = sim.PROGEMUP(data, runs=runs, output_dir=raw_dir)
    
    # Save results
    results_df.to_csv(f"{output_dir}/outputs.csv")
    print(f"Simulation complete. Results saved to {output_dir}/outputs.csv")
    
    # Generate visualizations
    try:
        create_plots(f"{output_dir}/outputs.csv", output_dir)
    except Exception as e:
        error_trace = traceback.format_exc()
        print(f"Warning: Failed to generate plots: {e}")
        print(error_trace)
        error_log_path = f"{output_dir}/plots_error.log"
        try:
            with open(error_log_path, "w") as f:
                f.write(f"Error generating plots:\n{error_trace}")
            print(f"Error details written to {error_log_path}")
        except Exception as log_error:
            print(f"Failed to write error log: {log_error}")
    
    # Generate Excel analysis report
    try:
        generate_excel_analysis(output_dir, f"{output_dir}/outputs.csv", raw_dir)
    except Exception as e:
        error_trace = traceback.format_exc()
        print(f"Warning: Failed to generate Excel analysis: {e}")
        print(error_trace)
    
    # Generate NET-specific analysis (add-on)
    try:
        from net_analysis import create_net_age_tier_plot, generate_net_summary
        
        # Enrich metadata for NET summary
        net_metadata = {
            **metadata,
            'master_seed': seed,
            'timestamp': timestamp
        }
        create_net_age_tier_plot(f"{output_dir}/outputs.csv", output_dir)
        generate_net_summary(f"{output_dir}/outputs.csv", raw_dir, net_metadata)
    except ImportError as e:
        print(f"Info: NET analysis module not available, skipping: {e}")
    except Exception as e:
        error_trace = traceback.format_exc()
        print(f"Warning: Failed to generate NET analysis: {e}")
        print(error_trace)
    
    print(f"\nBenchmark complete. All artifacts saved to: {output_dir}")
    return output_dir

if __name__ == "__main__":
    run_benchmark()
