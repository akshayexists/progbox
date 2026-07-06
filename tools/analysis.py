"""
analysis.py · Progbox, Monte Carlo Analysis

A diagnostic instrument for tuning progression scripts. The simulation
engine writes one row per (player, run) into outputs.csv; this module
reads that, joins it with the pre-simulation baseline ({RUN}/raw/input.csv),
and produces an interactive HTML dashboard plus an Excel workbook.

TWO MODES:

  SINGLE RUN     python analysis.py <run_dir>
                 The seven-section deep-dive dashboard (unchanged layout).

  COMPARISON     python analysis.py <run_dir_A> <run_dir_B> [...]
                 NEW. Overlays N progression scripts on the same axes and
                 emits a head-to-head SCORECARD of the tuning KPIs:
                   · peak (zero-crossing) age & post-peak decline slope
                   · prime-age talent separation (production terciles)
                   · league drift, run-to-run noise, ICC, Kendall's W
                   · ceiling behaviour (P99 OVR, % player-runs over cap)
                   · god-prog event rate
                 Labels are read from each run's metadata.json.

The dashboard is organized into seven thematic sections, each answering
a different class of tuning question:

    §1  LEAGUE HEALTH      Is the league drifting, inflating, churning?
    §2  AGE CURVE          When do players peak, how steep is decline?
    §3  ATTRIBUTE MOVEMENT Which ratings carry the curve, which lag?
    §4  PLAYER OUTCOMES    Risk/return per player, ceiling behaviour
    §5  INPUT SENSITIVITY  Do PER etc. actually drive outcomes?
    §6  RNG CALIBRATION    Signal-dominated or noise-dominated?
    §7  DIAGNOSTICS        Outliers, broken assumptions, heteroscedasticity

Column discovery is fully dynamic. Three buckets:

    META_COLS           Hardcoded in Config (bookkeeping columns).
    OVR_CALC_ORDER      Hardcoded in Config (the 15 OVR attributes).
                        Each is auto-classified as varying or fixed by
                        within-player variance across runs.
    MODEL_INPUTS        EVERYTHING ELSE numeric: PER, OBPM, DBPM, or
                        whatever new feature is added upstream.

HTML output is a single self-contained file (plotly.js inlined, no
network fetches) safe to embed via <iframe> / srcdoc. Charts are stored
as JSON and rendered lazily on scroll, so a 30-chart dashboard opens
instantly instead of rendering everything up front.

Usage:
    python analysis.py [run_dir ...] [--ceiling 84] [--full-excel] [--no-excel]
    from analysis import generate_analysis, generate_comparison
"""

from __future__ import annotations

import argparse
import json
import logging
import sys
import time
from pathlib import Path
from typing import Callable, Dict, List, Optional, Tuple

import numpy as np
import pandas as pd
import plotly.graph_objects as go
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from plotly.offline import get_plotlyjs
from plotly.subplots import make_subplots
from scipy import stats as scipy_stats
from scipy.optimize import curve_fit
from scipy.stats import gaussian_kde

# LOGGING
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
logger = logging.getLogger(__name__)


# CONFIG : Engine ground truths. NEVER edit between progression scripts.
class Config:
    """Engine constants that never change between simulation configs."""

    # OVR calculation order and coefficients (game engine formula)
    OVR_CALC_ORDER: List[str] = [
        "Hgt",
        "Str",
        "Spd",
        "Jmp",
        "End",
        "Ins",
        "Dnk",
        "FT",
        "3Pt",
        "oIQ",
        "dIQ",
        "Drb",
        "Pss",
        "2Pt",
        "Reb",
    ]
    OVR_COEFFS: np.ndarray = np.array(
        [
            0.159,
            0.0777,
            0.123,
            0.051,
            0.0632,
            0.0126,
            0.0286,
            0.0202,
            0.0726,
            0.133,
            0.159,
            0.059,
            0.062,
            0.01,
            0.01,
        ],
        dtype=float,
    )
    OVR_CENTERS: np.ndarray = np.array(
        [
            47.5,
            50.2,
            50.8,
            48.7,
            39.9,
            42.4,
            49.5,
            47.0,
            47.1,
            46.8,
            46.7,
            54.8,
            51.3,
            47.0,
            51.4,
        ],
        dtype=float,
    )

    # Non-attribute metadata columns excluded from stat discovery
    META_COLS: List[str] = [
        "Run",
        "RunSeed",
        "Name",
        "Team",
        "Age",
        "PlayerID",
        "Baseline",
        "Ovr",
        "Delta",
        "PctChange",
        "AboveBaseline",
    ]

    #  Analysis thresholds (tunable but not engine truths)
    MIN_KDE_SAMPLES: int = 10
    MIN_BIN_COUNT: int = 5
    MCSE_THRESHOLD: float = 0.5  # OVR pts, target convergence
    MAX_HEATMAP_ROWS: int = 40
    PHYSICAL_ATTRS = {"Spd", "Str", "Jmp", "End"}
    SHOOTING_ATTRS = {"Ins", "Dnk", "FT", "2Pt", "3Pt"}
    MENTAL_ATTRS = {"oIQ", "dIQ"}
    SKILL_ATTRS = {"Drb", "Pss", "Reb"}
    PHYSICAL_FIXED = {"Hgt"}

    #  Comparison-mode defaults
    CEILING_OVR: float = 84.0  # softCeil + ceilBand ("John Wall check")
    PRIME_AGES: Tuple[int, int] = (25, 27)
    SEP_BANDS: List[Tuple[int, int]] = [(25, 27), (28, 30), (31, 33), (34, 45)]
    KENDALL_MAX_RUNS: int = 200  # subsample runs for Kendall's W


# AESTHETICS
AGE_COLORS = {"Youngest": "#2563eb", "Middle": "#f97316", "Oldest": "#dc2626"}
Q_COLORS = {"Q1": "#dc2626", "Q2": "#f97316", "Q3": "#16a34a", "Q4": "#2563eb"}
GROUP_COLORS = {
    "Physical": "#dc2626",
    "Shooting": "#f97316",
    "Mental": "#2563eb",
    "Skill": "#16a34a",
    "Fixed": "#94a3b8",
}
# One distinct color per progression script in comparison mode
SCRIPT_PALETTE = [
    "#2563eb",
    "#dc2626",
    "#16a34a",
    "#f97316",
    "#7c3aed",
    "#0891b2",
    "#be185d",
    "#65a30d",
]
DIVERGING_CMAP = [
    "#0d47a1",
    "#42a5f5",
    "#e3f2fd",
    "#ffffff",
    "#fce4ec",
    "#ef5350",
    "#b71c1c",
]

_PLOTLY_TEMPLATE = go.layout.Template(
    layout=go.Layout(
        font=dict(
            family='Inter, -apple-system, BlinkMacSystemFont, "Segoe UI", '
            "Roboto, sans-serif",
            color="#1e293b",
        ),
        paper_bgcolor="#ffffff",
        plot_bgcolor="#f8fafc",
        margin=dict(l=80, r=40, t=90, b=80),
        hoverlabel=dict(
            bgcolor="rgba(255,255,255,0.96)",
            bordercolor="#cbd5e1",
            font_size=13,
            font_color="#0f172a",
            font_family="Inter, Roboto, sans-serif",
        ),
        xaxis=dict(
            gridcolor="#e2e8f0",
            zerolinecolor="#94a3b8",
            title=dict(standoff=15, font=dict(size=13, color="#334155")),
            tickfont=dict(size=11, color="#64748b"),
            automargin=True,
        ),
        yaxis=dict(
            gridcolor="#e2e8f0",
            zerolinecolor="#94a3b8",
            title=dict(standoff=15, font=dict(size=13, color="#334155")),
            tickfont=dict(size=11, color="#64748b"),
            automargin=True,
        ),
        legend=dict(
            bgcolor="rgba(255,255,255,0.85)",
            bordercolor="#e2e8f0",
            borderwidth=1,
            font=dict(size=12, color="#1e293b"),
        ),
        title=dict(font=dict(size=20, color="#0f172a"), x=0.05, pad=dict(t=15, b=15)),
    )
)


def hex_to_rgba(hex_color: str, alpha: float) -> str:
    h = hex_color.lstrip("#")
    r, g, b = (int(h[i : i + 2], 16) for i in (0, 2, 4))
    return f"rgba({r},{g},{b},{alpha})"


def _empty_fig(title: str, reason: str = "INSUFFICIENT DATA") -> go.Figure:
    fig = go.Figure()
    fig.update_layout(
        title=f"{title} · {reason}",
        template=_PLOTLY_TEMPLATE,
    )
    fig._is_placeholder = True  # explicit flag
    return fig


def _attr_group(attr: str) -> str:
    """Map an OVR attribute to its conceptual group."""
    if attr in Config.PHYSICAL_ATTRS:
        return "Physical"
    if attr in Config.PHYSICAL_FIXED:
        return "Fixed"
    if attr in Config.SHOOTING_ATTRS:
        return "Shooting"
    if attr in Config.MENTAL_ATTRS:
        return "Mental"
    if attr in Config.SKILL_ATTRS:
        return "Skill"
    return "Other"


# STATISTICS : stateless utilities
def zscore(s: pd.Series) -> pd.Series:
    """Z-score; returns zeros when std is zero (constant feature)."""
    std = s.std()
    if std == 0 or pd.isna(std):
        return pd.Series(0.0, index=s.index, dtype=float)
    return (s - s.mean()) / std


def bootstrap_ci(
    data: np.ndarray,
    statistic_fn: Callable = np.mean,
    n_bootstrap: int = 3000,
    confidence_level: float = 0.95,
    seed: int = 0,
) -> Tuple[float, float, float]:
    """Calculates a standard bootstrap confidence interval for a given statistic."""
    sample_size = len(data)

    # Handle edge cases for empty or single-element inputs
    if sample_size < 2:
        fallback_value = float(statistic_fn(data)) if sample_size else 0.0
        return fallback_value, fallback_value, fallback_value

    point_estimate = float(statistic_fn(data))

    # Generate random indices for resampling with replacement
    random_generator = np.random.default_rng(seed)
    resample_indices = random_generator.integers(
        0, sample_size, size=(n_bootstrap, sample_size)
    )

    # Calculate the statistic for each bootstrap sample
    if statistic_fn is np.mean:
        bootstrap_estimates = data[resample_indices].mean(
            axis=1
        )  # Vectorized fast path
    else:
        bootstrap_estimates = np.array(
            [statistic_fn(data[idx]) for idx in resample_indices]
        )

    # Calculate percentiles based on the confidence level
    alpha_tail = (1 - confidence_level) / 2
    lower_percentile = 100 * alpha_tail
    upper_percentile = 100 * (1 - alpha_tail)

    return (
        point_estimate,
        float(np.percentile(bootstrap_estimates, lower_percentile)),
        float(np.percentile(bootstrap_estimates, upper_percentile)),
    )


def cluster_bootstrap_ci(
    df: pd.DataFrame,
    value_col: str,
    cluster_col: str = "PlayerID",
    statistic_fn: Callable = np.mean,
    n_bootstrap: int = 1500,
    confidence_level: float = 0.95,
    seed: int = 0,
) -> Tuple[float, float, float]:
    """Cluster bootstrap CI accounting for within-cluster correlation.
    """
    n_rows = len(df)
    values = df[value_col].values
    clusters = df[cluster_col].values
    unique_clusters, inverse = np.unique(clusters, return_inverse=True)
    n_clusters = len(unique_clusters)

    if n_clusters < 2:
        fallback_value = float(statistic_fn(values)) if n_rows else 0.0
        return fallback_value, fallback_value, fallback_value

    point_estimate = float(statistic_fn(values))
    random_generator = np.random.default_rng(seed)
    resample_idx = random_generator.integers(
        0, n_clusters, size=(n_bootstrap, n_clusters)
    )

    if statistic_fn is np.mean:
        # Vectorized fast path: pooled mean of resampled clusters
        cluster_sums = np.bincount(inverse, weights=values, minlength=n_clusters)
        cluster_counts = np.bincount(inverse, minlength=n_clusters).astype(float)
        bootstrap_estimates = cluster_sums[resample_idx].sum(axis=1) / cluster_counts[
            resample_idx
        ].sum(axis=1)
    else:
        cluster_groups = {i: values[inverse == i] for i in range(n_clusters)}
        bootstrap_estimates = np.empty(n_bootstrap)
        for i in range(n_bootstrap):
            resampled = np.concatenate([cluster_groups[c] for c in resample_idx[i]])
            bootstrap_estimates[i] = statistic_fn(resampled)

    alpha_tail = (1 - confidence_level) / 2
    return (
        point_estimate,
        float(np.percentile(bootstrap_estimates, 100 * alpha_tail)),
        float(np.percentile(bootstrap_estimates, 100 * (1 - alpha_tail))),
    )


def cohens_d(g1: np.ndarray, g2: np.ndarray) -> float:
    """Hedges-corrected Cohen's d. Zero when undefined."""
    n1, n2 = len(g1), len(g2)
    if n1 + n2 < 3:
        return 0.0
    v1, v2 = np.var(g1, ddof=1), np.var(g2, ddof=1)
    sp = np.sqrt(((n1 - 1) * v1 + (n2 - 1) * v2) / (n1 + n2 - 2))
    if sp == 0:
        return 0.0
    d = (np.mean(g1) - np.mean(g2)) / sp
    return d * (1 - 3 / (4 * (n1 + n2) - 9))


def ols_fit(
    X: np.ndarray, y: np.ndarray, compute_se: bool = True
) -> Tuple[np.ndarray, Optional[np.ndarray], float]:
    """Stable QR-based OLS. Returns (beta, se, r_squared).

    Set compute_se=False to skip the SE computation (saves an O(p³)
    matrix inverse per call).
    """
    n, p = X.shape
    if n <= p:
        raise ValueError(f"Insufficient df: n={n} <= p={p}")
    Q, R = np.linalg.qr(X)
    beta = np.linalg.solve(R, Q.T @ y)
    resid = y - X @ beta
    rss = float(np.sum(resid**2))
    tss = float(np.sum((y - y.mean()) ** 2))
    r2 = 1.0 - rss / tss if tss > 0 else 0.0
    if not compute_se:
        return beta, None, r2
    sigma2 = rss / (n - p)
    R_inv = np.linalg.solve(R, np.eye(p))
    var_beta = sigma2 * (R_inv @ R_inv.T)
    se = np.sqrt(np.maximum(np.diag(var_beta), 0.0))
    return beta, se, r2


def partial_corr(x: np.ndarray, y: np.ndarray, Z: np.ndarray) -> Tuple[float, float]:
    """Partial Pearson r of x,y controlling for columns of Z.
    Returns (r, p). Z can be empty (returns plain Pearson)."""
    if Z.size == 0 or Z.shape[1] == 0:
        return scipy_stats.pearsonr(x, y)
    try:
        A = np.column_stack([np.ones(len(Z)), Z])
        Q, R = np.linalg.qr(A)
        bx = np.linalg.solve(R, Q.T @ x)
        by = np.linalg.solve(R, Q.T @ y)
        rx = x - A @ bx
        ry = y - A @ by
        return scipy_stats.pearsonr(rx, ry)
    except np.linalg.LinAlgError:
        return 0.0, 1.0


def mahalanobis_distance(X: np.ndarray) -> np.ndarray:
    """Mahalanobis distance from each row to the multivariate centroid."""
    cov = np.cov(X, rowvar=False)
    try:
        inv = np.linalg.inv(cov)
    except np.linalg.LinAlgError:
        inv = np.linalg.pinv(cov)
    d = X - X.mean(axis=0)
    return np.sqrt(np.einsum("ij,jk,ik->i", d, inv, d))


def kendalls_w(rank_matrix: np.ndarray) -> float:
    """Kendall's W coefficient of concordance across judges (= runs)."""
    n_items, n_judges = rank_matrix.shape
    if n_items < 2 or n_judges < 2:
        return np.nan
    R = rank_matrix.sum(axis=1)
    S = np.sum((R - R.mean()) ** 2)
    return 12 * S / (n_judges**2 * (n_items**3 - n_items))


def logistic(x: np.ndarray, b0: float, b1: float) -> np.ndarray:
    z = np.clip(b0 + b1 * x, -60.0, 60.0)
    return 1.0 / (1.0 + np.exp(-z))


def qcut_safe(s: pd.Series, q: int, labels=None) -> Optional[pd.Categorical]:
    """qcut that returns None instead of raising on degenerate input."""
    try:
        return pd.qcut(s.rank(method="first"), q, labels=labels)
    except ValueError:
        return None


# TUNING KPIs : shared between single-run charts and the comparison scorecard.
# These encode the diagnostics that matter for progression-script tuning:
# where the age curve crosses zero, how hard it declines, and whether
# same-age production actually separates outcomes at prime age.
def estimate_peak(per_age: pd.DataFrame) -> Dict[str, float]:
    """Interpolated zero-crossing (peak age) + post-peak decline slope.
    Returns dict(peak_age, decline_slope, decline_r2); NaN when undefined.
    """
    out = {"peak_age": np.nan, "decline_slope": np.nan, "decline_r2": np.nan}
    if per_age.empty or "Mean" not in per_age.columns:
        return out
    mean_vals = per_age["Mean"].values
    ages = per_age["Age"].values
    pos_to_neg = np.where(
        (np.sign(mean_vals[:-1]) > 0) & (np.sign(mean_vals[1:]) <= 0)
    )[0]
    if len(pos_to_neg) == 0:
        return out
    i = int(pos_to_neg[-1])
    m1, m2 = mean_vals[i], mean_vals[i + 1]
    a1, a2 = ages[i], ages[i + 1]
    out["peak_age"] = (
        float(a1 + (0 - m1) * (a2 - a1) / (m2 - m1)) if m2 != m1 else float(a1)
    )
    post = per_age.iloc[i + 1 :]
    if len(post) >= 3:
        res = scipy_stats.linregress(post["Age"], post["Mean"])
        out["decline_slope"] = float(res.slope)
        out["decline_r2"] = float(res.rvalue**2)
    return out


def production_series(ppl: pd.DataFrame) -> Tuple[Optional[pd.Series], str]:
    """Best available production composite for separation KPIs."""
    if "OBPM" in ppl.columns and "DBPM" in ppl.columns:
        return ppl["OBPM"] + ppl["DBPM"], "OBPM+DBPM"
    if "PER" in ppl.columns:
        return ppl["PER"], "PER"
    return None, ""


def separation_gap(ppl: pd.DataFrame, lo: int, hi: int) -> Tuple[float, int]:
    """Mean-delta gap: top vs bottom production tercile within an age band.

    at prime age, do studs actually out-progress scrubs, or does everyone move together?
    Returns (gap, n_players); (nan, n) when undefined."""
    prod, _ = production_series(ppl)
    if prod is None:
        return np.nan, 0
    sub = ppl[(ppl["Age"] >= lo) & (ppl["Age"] <= hi)]
    if len(sub) < 9:
        return np.nan, len(sub)
    q = qcut_safe(prod.loc[sub.index], 3, labels=["lo", "mid", "hi"])
    if q is None:
        return np.nan, len(sub)
    gap = (
        sub.loc[q == "hi", "MeanDelta"].mean() - sub.loc[q == "lo", "MeanDelta"].mean()
    )
    return float(gap), len(sub)


def partial_separation_gap(
    ppl: pd.DataFrame, lo: int, hi: int, ovr_bin: float = 8.0
) -> Tuple[float, int]:
    """Production-tercile ΔOVR gap WITHIN baseline-OVR bands, then averaged.

    This variant controls for baseline OVR by
    computing the tercile gap inside fixed OVR bands and averaging them. Essentially,
    'at the SAME OVR, does talent still separate?' Positive = yes. Returns (gap, n_players)."""
    prod, _ = production_series(ppl)
    if prod is None:
        return np.nan, 0
    sub = ppl[(ppl["Age"] >= lo) & (ppl["Age"] <= hi)].dropna(subset=["Baseline"])
    if len(sub) < 18:
        return np.nan, len(sub)
    edges = np.arange(sub["Baseline"].min(), sub["Baseline"].max() + ovr_bin, ovr_bin)
    gaps = []
    for b0 in edges:
        band = sub[(sub["Baseline"] >= b0) & (sub["Baseline"] < b0 + ovr_bin)]
        if len(band) < 9:
            continue
        q = qcut_safe(prod.loc[band.index], 3, labels=["lo", "mid", "hi"])
        if q is None:
            continue
        g = (
            band.loc[q == "hi", "MeanDelta"].mean()
            - band.loc[q == "lo", "MeanDelta"].mean()
        )
        if not np.isnan(g):
            gaps.append(g)
    return (float(np.mean(gaps)) if gaps else np.nan), len(sub)


# COLUMN REGISTRY : Dynamic discovery of column roles.
class ColumnRegistry:
    """Classifies every numeric column in the simulation output.

    Four disjoint buckets:

      meta            Bookkeeping (Config.META_COLS): Run, PlayerID, Delta, ...
      varying_attrs   OVR_CALC_ORDER attrs that change across runs
      fixed_attrs     OVR_CALC_ORDER attrs that don't change (e.g. Hgt)
      inputs          EVERYTHING ELSE are considered model inputs (PER, OBPM, ...)

    A column with sub-1e-3 within-player std is considered fixed.
    A non-OVR column with zero overall std is dropped (useless feature).
    """

    def __init__(self, df: pd.DataFrame):
        numeric = df.select_dtypes(include=[np.number]).columns.tolist()
        self.meta: List[str] = [c for c in Config.META_COLS if c in df.columns]
        self.varying_attrs: List[str] = []
        self.fixed_attrs: List[str] = []
        self.inputs: List[str] = []

        candidates = [c for c in numeric if c not in self.meta]
        if candidates:
            within_all = df.groupby("PlayerID")[candidates].std().mean()
            overall_std = df[candidates].std()
        for col in candidates:
            varies = within_all[col] > 1e-3
            if col in Config.OVR_CALC_ORDER:
                (self.varying_attrs if varies else self.fixed_attrs).append(col)
            else:
                if overall_std[col] > 1e-6:
                    self.inputs.append(col)

        # Preserve OVR_CALC_ORDER for varying/fixed (more readable charts)
        order_map = {a: i for i, a in enumerate(Config.OVR_CALC_ORDER)}
        self.varying_attrs.sort(key=lambda a: order_map[a])
        self.fixed_attrs.sort(key=lambda a: order_map[a])

        logger.info(
            f"Column registry  ::  {len(self.varying_attrs)} varying attrs  ·  "
            f"{len(self.fixed_attrs)} fixed attrs  ·  "
            f"{len(self.inputs)} model inputs"
        )
        logger.info(f"  varying  :: {self.varying_attrs}")
        if self.fixed_attrs:
            logger.info(f"  fixed    :: {self.fixed_attrs}")
        logger.info(f"  inputs   :: {self.inputs}")

    @property
    def all_predictors(self) -> List[str]:
        """Predictor set for sensitivity analyses: model inputs + Age + Baseline."""
        base = ["Age", "Baseline"] if "Baseline" in self.meta else ["Age"]
        return base + self.inputs


# DATA LOADER : Reads simulation outputs and baseline from disk.
class DataLoader:
    def __init__(self, base: Path = Path("../outputs/")):
        self._base = base

    def find_latest_run(self) -> Path:
        """Most recently modified run directory under base."""
        hits = sorted(
            self._base.glob("*/raw/outputs.csv"),
            key=lambda p: p.stat().st_mtime,
            reverse=True,
        )
        if not hits:
            raise FileNotFoundError(f"No simulation outputs found under {self._base}/")
        run_dir = hits[0].parent.parent
        logger.info(f"Auto-discovered run: {run_dir}")
        return run_dir

    @staticmethod
    def load_sim(raw_dir: Path) -> pd.DataFrame:
        path = raw_dir / "outputs.csv"
        if not path.exists():
            raise FileNotFoundError(f"Missing: {path}")
        df = pd.read_csv(path)
        logger.info(f"Loaded {len(df):,} simulation rows from {path}")
        return df

    @staticmethod
    def load_baseline(raw_dir: Path) -> pd.DataFrame:
        path = raw_dir / "input.csv"
        if not path.exists():
            logger.error(f"CRITICAL: {path} missing. attribute deltas will be skipped.")
            return pd.DataFrame()
        df = pd.read_csv(path, index_col=0)
        df.index.name = "PlayerID"
        logger.info(f"Loaded baseline for {len(df)} players from {path}")
        return df

    @staticmethod
    def load_metadata(run_dir: Path) -> Dict:
        """metadata.json written by the C++ orchestrator; {} when absent."""
        path = run_dir / "metadata.json"
        if not path.exists():
            return {}
        try:
            return json.loads(path.read_text())
        except (json.JSONDecodeError, OSError):
            logger.warning(f"Unreadable metadata.json in {run_dir}")
            return {}

    @staticmethod
    def load_godprogs(run_dir: Path) -> int:
        """Number of god-prog events recorded for this run set; 0 when absent."""
        path = run_dir / "raw" / "godprogs.json"
        if not path.exists():
            return 0
        try:
            return len(json.loads(path.read_text(encoding="utf8")))
        except (json.JSONDecodeError, OSError):
            return 0


# DATASET : Raw data + all derived computations, lazily cached.
class Dataset:
    """Holds the loaded simulation output plus every cached derivation.

    All heavy lifting is done once per property and memoized. Downstream
    chart builders read from these properties rather than recomputing.
    """

    def __init__(self, sim: pd.DataFrame, baseline: pd.DataFrame):
        self.sim = sim.copy()
        self.baseline = baseline
        self.cols = ColumnRegistry(self.sim)

        # Stamp the age tier once at load time
        self.sim["AgeTier"] = self.assign_tiers(self.sim["Age"])

        # Caches
        self._per_player: Optional[pd.DataFrame] = None
        self._attr_deltas: Optional[pd.DataFrame] = None
        self._per_age: Optional[pd.DataFrame] = None
        self._icc: Optional[pd.DataFrame] = None
        self._convergence: Optional[Dict] = None
        self._top_drivers: Optional[List[str]] = None

    #  helpers

    @staticmethod
    def assign_tiers(ages: pd.Series) -> pd.Series:
        """Bin ages into 3 quantile tiers, gracefully degrade with few values."""
        if ages.nunique() < 3:
            return pd.Series(["Middle"] * len(ages), index=ages.index)
        try:
            return pd.qcut(ages, q=3, labels=list(AGE_COLORS.keys()), duplicates="drop")
        except ValueError:
            return pd.Series(["Middle"] * len(ages), index=ages.index)

    #  per-player summary
    @property
    def per_player(self) -> pd.DataFrame:
        """Computes a comprehensive per-player summary DataFrame from the underlying simulation data.        """
        if self._per_player is not None:
            return self._per_player

        player_groupby = self.sim.groupby("PlayerID", sort=True)

        aggregation_mapping = {
            "Name": ("Name", "first"),
            "Team": ("Team", "first"),
            "Age": ("Age", "first"),
            "Baseline": ("Baseline", "first"),
            "MeanDelta": ("Delta", "mean"),
            "StdDelta": ("Delta", "std"),
            "P50_Delta": ("Delta", "median"),
            "MinDelta": ("Delta", "min"),
            "MaxDelta": ("Delta", "max"),
            "MeanOvr": ("Ovr", "mean"),
            "StdOvr": ("Ovr", "std"),
        }

        # Include additional simulation configuration inputs if not already mapped
        for input_column in self.cols.inputs:
            if input_column not in aggregation_mapping:
                aggregation_mapping[input_column] = (input_column, "first")

        summary_df = player_groupby.agg(**aggregation_mapping).reset_index()

        # 2. Quantiles via vectorized calculations on the grouped Series
        delta_groupby_series = player_groupby["Delta"]

        target_quantiles = [
            (0.05, "P05_Delta"),
            (0.25, "P25_Delta"),
            (0.75, "P75_Delta"),
            (0.95, "P95_Delta"),
        ]
        for quantile_value, column_label in target_quantiles:
            summary_df[column_label] = delta_groupby_series.quantile(
                quantile_value
            ).values

        # PERF: fully vectorized (mean of a boolean column) instead of
        # groupby.apply with a Python lambda per player.
        summary_df["PctPositive"] = (
            (self.sim["Delta"] > 0)
            .groupby(self.sim["PlayerID"])
            .mean()
            .sort_index()
            .values
        )
        summary_df["AgeTier"] = self.assign_tiers(summary_df["Age"])

        # Vectorized string concatenation to replace row-by-row df.apply(axis=1)
        team_or_free_agent = summary_df["Team"].fillna("FA")
        summary_df["Label"] = summary_df["Name"] + " (" + team_or_free_agent + ")"

        # Spread Metrics
        summary_df["IQR"] = summary_df["P75_Delta"] - summary_df["P25_Delta"]
        summary_df["Range90"] = summary_df["P95_Delta"] - summary_df["P05_Delta"]

        # Risk-Adjusted Return Metrics
        summary_df["Sharpe"] = np.where(
            summary_df["StdDelta"] > 0,
            summary_df["MeanDelta"] / summary_df["StdDelta"],
            0.0,
        )
        # Certainty-Equivalent (CARA utility function, Risk Aversion Parameter A=1): μ − ½σ²
        summary_df["CertEquiv"] = summary_df["MeanDelta"] - 0.5 * (
            summary_df["StdDelta"] ** 2
        )

        # 4. Age-Tier-Adjusted Performance Z-Scores (Performance relative to cohort)
        age_tier_stats = (
            summary_df.groupby("AgeTier", observed=True)["MeanDelta"]
            .agg(["mean", "std"])
            .rename(columns={"mean": "cohort_mean", "std": "cohort_std"})
        )

        summary_df = summary_df.merge(
            age_tier_stats, left_on="AgeTier", right_index=True, how="left"
        )

        summary_df["AgeAdjZ"] = np.where(
            summary_df["cohort_std"] > 0,
            (summary_df["MeanDelta"] - summary_df["cohort_mean"])
            / summary_df["cohort_std"],
            0.0,
        )

        # Drop temporary calculation columns
        summary_df = summary_df.drop(columns=["cohort_mean", "cohort_std"])

        self._per_player = summary_df
        return summary_df

    #  per-age summary
    @property
    def per_age(self) -> pd.DataFrame:
        """One row per integer age."""
        if self._per_age is not None:
            return self._per_age

        # Use SIM rows (not per-player) so each (player, run) contributes
        rows = []
        for age, sub in self.sim.groupby("Age"):
            sub_clean = sub[["PlayerID", "Delta"]].dropna()
            if len(sub_clean) < Config.MIN_BIN_COUNT:
                continue
            mean, lo, hi = cluster_bootstrap_ci(
                sub_clean, "Delta", n_bootstrap=1500, seed=int(age) * 31
            )
            d = sub_clean["Delta"].values
            rows.append(
                {
                    "Age": int(age),
                    "N": len(d),
                    "NPlayers": sub_clean["PlayerID"].nunique(),
                    "Mean": mean,
                    "CI_lo": lo,
                    "CI_hi": hi,
                    "Std": d.std(),
                    "P05": np.percentile(d, 5),
                    "P50": np.percentile(d, 50),
                    "P95": np.percentile(d, 95),
                    "PPos": float((d > 0).mean()),
                }
            )
        self._per_age = pd.DataFrame(rows).sort_values("Age").reset_index(drop=True)
        return self._per_age

    #  attribute deltas (per-player mean attribute movement)
    @property
    def attr_deltas(self) -> pd.DataFrame:
        """
        Per-player mean simulated attribute - baseline attribute.
        Returns an empty frame if baseline data isn't available.
        """
        if self._attr_deltas is not None:
            return self._attr_deltas
        if self.baseline.empty:
            self._attr_deltas = pd.DataFrame()
            return self._attr_deltas

        avail = [
            c
            for c in self.cols.varying_attrs
            if c in self.baseline.columns and c in self.sim.columns
        ]
        mean_sim = self.sim.groupby("PlayerID")[avail].mean()
        shared = mean_sim.index.intersection(self.baseline.index)
        delta = mean_sim.loc[shared] - self.baseline[avail].loc[shared]
        delta = delta.join(self.sim.groupby("PlayerID")["Age"].first())
        self._attr_deltas = delta
        return delta

    #  ICC variance decomposition
    @property
    def icc(self) -> pd.DataFrame:
        """Intraclass correlation per varying attribute: σ²_between / total.

        High ICC --> outcomes determined by player identity (low RNG influence).
        Low  ICC --> outcomes dominated by run-to-run noise.
        """
        if self._icc is not None:
            return self._icc
        rows = []
        for attr in self.cols.varying_attrs:
            if attr not in self.sim.columns:
                continue
            grp = self.sim.groupby("PlayerID")[attr]
            between = grp.mean().var()
            within = grp.var().mean()
            total = between + within
            rows.append(
                {
                    "Attribute": attr,
                    "Group": _attr_group(attr),
                    "Var_Between": between,
                    "Var_Within": within,
                    "Total_Var": total,
                    "ICC": between / total if total > 0 else np.nan,
                }
            )
        # Also include OVR for context
        grp = self.sim.groupby("PlayerID")["Ovr"]
        b, w = grp.mean().var(), grp.var().mean()
        rows.append(
            {
                "Attribute": "Ovr (agg)",
                "Group": "Aggregate",
                "Var_Between": b,
                "Var_Within": w,
                "Total_Var": b + w,
                "ICC": b / (b + w) if (b + w) > 0 else np.nan,
            }
        )
        self._icc = pd.DataFrame(rows).sort_values("ICC", ascending=True)
        return self._icc

    #  convergence
    @property
    def convergence(self) -> Dict:
        """MCSE diagnostics. Tells you if you ran enough Monte Carlo passes."""
        if self._convergence is not None:
            return self._convergence
        grp = self.sim.groupby("PlayerID")["Delta"]
        s = grp.agg(Mean="mean", Std="std", N="count")
        s["MCSE"] = s["Std"] / np.sqrt(s["N"])
        s["MCSE"] = s["MCSE"].replace([np.inf, np.nan], 0.0)
        self._convergence = {
            "max_mcse": float(s["MCSE"].max()),
            "mean_mcse": float(s["MCSE"].mean()),
            "pct_converged": float((s["MCSE"] < Config.MCSE_THRESHOLD).mean()),
            "top_volatile": s.nlargest(6, "Std").index.tolist(),
            "stats": s,
        }
        return self._convergence

    #  top drivers among model inputs
    @property
    def top_drivers(self) -> List[str]:
        """Top 3 model inputs (incl. Age, Baseline) by |partial r| vs MeanDelta."""
        if self._top_drivers is not None:
            return self._top_drivers
        ppl = self.per_player
        ctrl = ppl[["Age"]].values  # always control for Age
        scored = []
        for col in self.cols.all_predictors:
            if col == "Age" or col not in ppl.columns or ppl[col].std() == 0:
                continue
            other_ctrl = ctrl if col != "Age" else np.array([]).reshape(len(ppl), 0)
            r, _ = partial_corr(ppl[col].values, ppl["MeanDelta"].values, other_ctrl)
            if not np.isnan(r):
                scored.append((col, abs(r), r))
        scored.sort(key=lambda x: x[1], reverse=True)
        # Always pin Age at the front
        result = ["Age"] + [s[0] for s in scored if s[0] != "Age"][:2]
        self._top_drivers = result[:3]
        logger.info(f"Top drivers: {self._top_drivers}")
        return self._top_drivers


# CHART BUILDER : One method per chart, organized by HTML section.
#
# Every method returns a single go.Figure. Methods are stateless beyond
# reading `self._ds`. Each docstring leads with the TUNING TAKEAWAY that
# the chart provides, i.e. what to look for when deciding whether the
# progression script needs adjustment.
# Section spec: title, intro text, list of chart-method names.
SECTIONS: List[Dict] = [
    {
        "id": "league-health",
        "title": "§1 · League Health",
        "intro": "Where the league as a whole has moved. If the "
        "population mean drifts strongly in either direction "
        "across runs, the script has a systemic bias. Look for: "
        "stable mean OVR, comparable spread before/after, no "
        "collapse of the top cohort, and a near-diagonal "
        "quintile transition matrix.",
        "charts": [
            "chart_pre_vs_post_ovr",
            "chart_pre_post_scatter",
            "chart_cohort_transition",
        ],
    },
    {
        "id": "age-curve",
        "title": "§2 · Age Curve",
        "intro": "Probably the script's most consequential property."
        "The per-group view "
        "checks whether Physical / Shooting / Mental / Skill "
        "follow different lifecycles, as they should.",
        "charts": [
            "chart_age_curve",
            "chart_age_curve_bands",
            "chart_age_group_curves",
        ],
    },
    {
        "id": "attribute-movement",
        "title": "§3 · Attribute Movement",
        "intro": "Identify which ratings drive the curve versus those that are inert. "
        "Start with the Age x Attribute heatmap for the macro picture, "
        "then drill into per-attribute panels. The two ΔOVR decompositions "
        "translate raw movement into OVR impact: one splits the LEVEL of "
        "ΔOVR by attribute, the other splits the BETWEEN-PLAYER SPREAD of "
        "ΔOVR (why some players move more than others). Co-movement splits "
        "into between-player (tuning) and within-run (RNG) correlations, "
        "while saturation reveals clamp behavior.",
        "charts": [
            "chart_attr_age_heatmap",
            "chart_per_attribute_age_curves",
            "chart_ovr_decomposition",
            "chart_ovr_variance_decomposition",
            "chart_attr_comovement",
            "chart_attribute_saturation",
        ],
    },
    {
        "id": "player-outcomes",
        "title": "§4 · Per-Player Outcomes",
        "intro": "Individual risk/return. Look for a populated 'star quadrant' "
        "(high mean gain, low spread) and a sparse 'danger quadrant' "
        "(low mean, high spread). The improvement-probability curve"
        "shows the shape of your model's distribution."
        "The horsetail plot exposes per-player tail asymmetry.",
        "charts": [
            "chart_risk_return",
            "chart_outcome_distributions",
            "chart_improve_probability",
        ],
    },
    {
        "id": "input-sensitivity",
        "title": "§5 · Input Sensitivity",
        "intro": "Evaluate whether model inputs (PER, OBPM, DBPM, ...) steer outcomes "
        "Controlled effects (holding Age and Baseline constant) "
        "are the headline metrics. Shapley R² provides stable importance rankings, "
        "residualized partial dependence reveals "
        "the functional shape (linear, threshold, or non-monotonic) of each effect.",
        "charts": [
            "chart_controlled_input_effects",
            "chart_incremental_r2",
            "chart_partial_dependence",
        ],
    },
    {
        "id": "rng-calibration",
        "title": "§6 · RNG Calibration",
        "intro": "Measure outcome variance: signal (player identity) vs. noise (RNG). "
        "Aim for ICC > 0.4 on most varying attributes. Below 0.4 is too random; "
        "above 0.9 is deterministic, so your script is extra deterministic. Convergence "
        "verifies if you ran enough passes, while rank stability shows if "
        "run-to-run ordering is too noisy or too rigid.",
        "charts": [
            "chart_icc",
            "chart_convergence",
            "chart_rank_stability",
        ],
    },
    {
        "id": "diagnostics",
        "title": "§7 · Diagnostics",
        "intro": "Identify where script assumptions break. Outliers flag individuals with "
        "highly improbable joint (μΔ, σΔ) signatures. The funnel plot reveals "
        "if variance collapses near an OVR ceiling.",
        "charts": [
            "chart_outlier_detection",
            "chart_funnel",
        ],
    },
    {
        "id": "player-explorer",
        "title": "§8 · Player Explorer",
        "intro": "Any single player's full outcome, on demand. Pick a player "
        "(type to search and hit enter) to see the distribution of their OVR Δ over "
        "all runs, with their mean, P5/P95, and the league mean for "
        "reference. ",
        "charts": ["player_explorer"],
    },
]

# Charts that render as self-contained HTML widgets (a JSON payload + their
# own script), not as go.Figure. build_all and the registry validator skip
# them; they're injected into the charts dict by the pipeline.
WIDGET_CHARTS = {"player_explorer"}


class ChartBuilder:
    """Generates every Plotly figure consumed by the HTML dashboard."""

    def __init__(self, ds: Dataset):
        self._ds = ds

    # Convenience accessors
    @property
    def sim(self) -> pd.DataFrame:
        return self._ds.sim

    @property
    def ppl(self) -> pd.DataFrame:
        return self._ds.per_player

    @property
    def ad(self) -> pd.DataFrame:
        return self._ds.attr_deltas

    @property
    def cols(self) -> ColumnRegistry:
        return self._ds.cols

    #  Build orchestration

    def build_all(self) -> Dict[str, go.Figure]:
        """Build every chart for every section. Returns dict keyed by 'sec/key'."""
        out: Dict[str, go.Figure] = {}
        total = sum(
            1 for sec in SECTIONS for c in sec["charts"] if c not in WIDGET_CHARTS
        )
        i = 0
        for sec in SECTIONS:
            for method_name in sec["charts"]:
                if method_name in WIDGET_CHARTS:
                    continue  # widgets are injected by the pipeline, not built here
                i += 1
                t0 = time.time()
                key = f"{sec['id']}/{method_name}"
                method = getattr(self, method_name)
                logger.info(f"  [{i:>2}/{total}] {method_name}")
                try:
                    fig = method()
                    dt = time.time() - t0
                    status = f"{dt:.2f}s"
                    logger.info(f"           -> {status}")
                except Exception as e:
                    logger.error(
                        f"           -> FAILED: {type(e).__name__}: {e}",
                        exc_info=True,
                    )
                    fig = _empty_fig(method_name, f"ERROR: {type(e).__name__}")
                out[key] = fig
        return out

    # ════════════════════════════════════════════════════════════════════
    # §1  LEAGUE HEALTH
    # ════════════════════════════════════════════════════════════════════

    def chart_pre_vs_post_ovr(self) -> go.Figure:
        """Pre- vs post-simulation OVR distributions overlaid as KDEs.

        TUNING TAKEAWAY: If the post curve is shifted right of pre, the
        league is inflating; if shifted left, deflating. Width changes
        reveal whether the script compresses or spreads outcomes.

        STATS NOTE: pre and post are the SAME players (a paired sample,
        r≈0.95 here), so a two-sample KS/independent Cohen's d is invalid as
        it ignores the pairing and badly under-powers the test. We report
        the paired Wilcoxon signed-rank test, the paired effect size
        d_z = mean(Δ)/sd(Δ), and a bootstrap CI on the mean shift. The KDE
        overlay is purely a visual of the two marginals.
        """
        # Use rows with BOTH values so the pairing is intact.
        paired = self.ppl.dropna(subset=["Baseline", "MeanOvr"])
        pre = paired["Baseline"].values
        post = paired["MeanOvr"].values
        if len(pre) < Config.MIN_KDE_SAMPLES:
            return _empty_fig("OVR Distribution: Pre vs Post")

        x_grid = np.linspace(
            min(pre.min(), post.min()) - 3, max(pre.max(), post.max()) + 3, 400
        )
        kde_pre = gaussian_kde(pre, bw_method="scott")(x_grid)
        kde_post = gaussian_kde(post, bw_method="scott")(x_grid)

        fig = go.Figure()
        fig.add_trace(
            go.Scatter(
                x=x_grid,
                y=kde_pre,
                name="Pre-Simulation (baseline)",
                mode="lines",
                fill="tozeroy",
                fillcolor=hex_to_rgba("#94a3b8", 0.25),
                line=dict(color="#475569", width=2, shape="spline"),
                hovertemplate="Pre  |  OVR %{x:.1f}<br>Density %{y:.4f}<extra></extra>",
            )
        )
        fig.add_trace(
            go.Scatter(
                x=x_grid,
                y=kde_post,
                name="Post-Simulation (mean)",
                mode="lines",
                fill="tozeroy",
                fillcolor=hex_to_rgba("#2563eb", 0.20),
                line=dict(color="#2563eb", width=3, shape="spline"),
                hovertemplate="Post  |  OVR %{x:.1f}<br>Density %{y:.4f}<extra></extra>",
            )
        )
        for v, c, lbl in [
            (pre.mean(), "#475569", "Pre μ"),
            (post.mean(), "#2563eb", "Post μ"),
        ]:
            fig.add_vline(
                x=v,
                line_dash="dash",
                line_color=c,
                annotation_text=f"{lbl}={v:.1f}",
                annotation_position="top",
                annotation_font_color=c,
            )

        # Paired statistics (correct for same-player pre/post).
        diff = post - pre
        shift = float(diff.mean())
        _, lo, hi = bootstrap_ci(diff, n_bootstrap=3000, seed=7)
        sd_diff = diff.std(ddof=1)
        d_z = shift / sd_diff if sd_diff > 0 else 0.0
        try:
            if np.any(diff != 0):
                _, p_w = scipy_stats.wilcoxon(diff)
            else:
                p_w = 1.0
        except ValueError:
            p_w = float("nan")
        comp = ((post.std() / pre.std()) - 1) * 100 if pre.std() > 0 else 0.0

        mag = (
            "negligible"
            if abs(d_z) < 0.2
            else "small"
            if abs(d_z) < 0.5
            else "medium"
            if abs(d_z) < 0.8
            else "large"
        )
        direction = "inflating" if shift > 0 else "deflating" if shift < 0 else "flat"
        sig = "significant" if (p_w == p_w and p_w < 0.05) else "n.s."
        fig.add_annotation(
            text=(
                f"<b>League {direction}</b>: {shift:+.2f} OVR "
                f"[95% CI {lo:+.2f}, {hi:+.2f}]<br>"
                f"<b>Paired effect</b>: d_z={d_z:+.2f} ({mag})<br>"
                f"<b>Wilcoxon</b>: p={p_w:.2e} ({sig})<br>"
                f"<b>Spread change</b>: {comp:+.1f}%"
            ),
            xref="paper",
            yref="paper",
            x=0.98,
            y=0.98,
            showarrow=False,
            align="right",
            font=dict(size=11, family="monospace", color="#0f172a"),
            bgcolor="rgba(255,255,255,0.92)",
            bordercolor="#e2e8f0",
            borderwidth=1,
        )
        fig.update_layout(
            template=_PLOTLY_TEMPLATE,
            title="OVR Distribution :: Pre vs Post Simulation",
            xaxis_title="OVR",
            yaxis_title="Probability Density",
            hovermode="x unified",
        )
        return fig

    def chart_pre_post_scatter(self) -> go.Figure:
        """Per-player pre-->post OVR scatter against the y=x identity line.

        TUNING TAKEAWAY: Points above y=x improved, below regressed. The
        slope of the trend tells you whether the script is regressive
        (slope < 1, compresses talent) or expansive (slope > 1).
        """
        ppl = self.ppl.dropna(subset=["Baseline", "MeanOvr"])
        if len(ppl) < 8:
            return _empty_fig("Pre-->Post OVR Scatter")

        fig = go.Figure()
        for tier, color in AGE_COLORS.items():
            sub = ppl[ppl["AgeTier"] == tier]
            if sub.empty:
                continue
            fig.add_trace(
                go.Scatter(
                    x=sub["Baseline"],
                    y=sub["MeanOvr"],
                    mode="markers",
                    name=tier,
                    marker=dict(
                        color=color,
                        size=8,
                        line=dict(width=1, color="white"),
                        opacity=0.85,
                    ),
                    customdata=sub[["Label", "Age", "MeanDelta", "StdDelta"]].values,
                    hovertemplate=(
                        "<b>%{customdata[0]}</b><br>"
                        "Age %{customdata[1]}  |  Pre %{x:.1f} --> Post %{y:.1f}<br>"
                        "Δ %{customdata[2]:+.1f}  ±  %{customdata[3]:.1f}<extra></extra>"
                    ),
                )
            )

        # y=x reference and OLS slope
        lo, hi = ppl["Baseline"].min() - 2, ppl["Baseline"].max() + 2
        fig.add_trace(
            go.Scatter(
                x=[lo, hi],
                y=[lo, hi],
                mode="lines",
                name="No change (y=x)",
                line=dict(color="#94a3b8", width=2, dash="dot"),
                hoverinfo="skip",
            )
        )
        try:
            res = scipy_stats.linregress(ppl["Baseline"], ppl["MeanOvr"])
            xfit = np.array([lo, hi])
            yfit = res.intercept + res.slope * xfit
            fig.add_trace(
                go.Scatter(
                    x=xfit,
                    y=yfit,
                    mode="lines",
                    name=f"OLS slope={res.slope:.3f}",
                    line=dict(color="#dc2626", width=2, dash="dash"),
                    hoverinfo="skip",
                )
            )
            interp = (
                "compressive"
                if res.slope < 0.95
                else "expansive"
                if res.slope > 1.05
                else "neutral"
            )
            fig.add_annotation(
                text=(
                    f"<b>OLS</b>: y = {res.intercept:+.2f} + {res.slope:.3f}·x<br>"
                    f"<b>R²</b> = {res.rvalue**2:.3f}<br>"
                    f"<b>Regime</b>: {interp}"
                ),
                xref="paper",
                yref="paper",
                x=0.02,
                y=0.98,
                showarrow=False,
                align="left",
                font=dict(size=11, family="monospace", color="#0f172a"),
                bgcolor="rgba(255,255,255,0.92)",
                bordercolor="#e2e8f0",
                borderwidth=1,
            )
        except ValueError:
            pass

        fig.update_layout(
            template=_PLOTLY_TEMPLATE,
            title="Per-Player Pre --> Post OVR",
            xaxis_title="Baseline OVR",
            yaxis_title="Mean Simulated OVR",
            hovermode="closest",
        )
        return fig

    def chart_cohort_transition(self) -> go.Figure:
        """Pre-sim quintile x post-sim quintile transition matrix."""
        ppl = self.ppl.dropna(subset=["Baseline", "MeanOvr"])
        if len(ppl) < 25:
            return _empty_fig("Cohort Quintile Transition")

        labels = ["Q1 (low)", "Q2", "Q3", "Q4", "Q5 (high)"]
        q_pre = pd.qcut(ppl["Baseline"].rank(method="first"), 5, labels=labels)
        q_post = pd.qcut(ppl["MeanOvr"].rank(method="first"), 5, labels=labels)
        mat = pd.crosstab(q_pre, q_post, normalize="index")

        fig = go.Figure(
            data=go.Heatmap(
                z=mat.values,
                x=mat.columns.astype(str),
                y=mat.index.astype(str),
                colorscale="Blues",
                zmin=0,
                zmax=1,
                text=(mat * 100).round(0).astype(int).astype(str) + "%",
                texttemplate="%{text}",
                textfont=dict(size=13),
                hovertemplate="From %{y} --> %{x}<br>%{z:.1%} of row<extra></extra>",
                xgap=3,
                ygap=3,
                colorbar=dict(title="Row %", thickness=15, tickformat=".0%"),
            )
        )
        diag = float(np.diag(mat).mean())
        fig.update_layout(
            template=_PLOTLY_TEMPLATE,
            title=(
                f"Quintile Transition (Pre --> Post) :: avg diagonal retention {diag:.0%}"
            ),
            xaxis_title="Post-Sim Quintile",
            yaxis_title="Pre-Sim Quintile",
            yaxis_autorange="reversed",
            height=520,
        )
        return fig

    # ════════════════════════════════════════════════════════════════════
    # §2  AGE CURVE
    # ════════════════════════════════════════════════════════════════════

    def chart_age_curve(self) -> go.Figure:
        """Mean OVR Δ per integer age with bootstrap 95% CI band.

        TUNING TAKEAWAY: This is THE chart. Mark where the curve crosses
        zero (the peak age), where the slope inflects (decline onset),
        and how steep the decline is. Compare against intent.
        """
        pa = self._ds.per_age
        if pa.empty:
            return _empty_fig("Age Curve")

        fig = go.Figure()
        # CI ribbon
        fig.add_trace(
            go.Scatter(
                x=pd.concat([pa["Age"], pa["Age"][::-1]]),
                y=pd.concat([pa["CI_hi"], pa["CI_lo"][::-1]]),
                fill="toself",
                fillcolor=hex_to_rgba("#2563eb", 0.15),
                line=dict(color="rgba(0,0,0,0)"),
                name="95% CI",
                hoverinfo="skip",
            )
        )
        # Mean curve
        fig.add_trace(
            go.Scatter(
                x=pa["Age"],
                y=pa["Mean"],
                mode="lines+markers",
                line=dict(color="#2563eb", width=3),
                marker=dict(size=8, color="#2563eb"),
                customdata=pa[["N", "NPlayers", "PPos", "Std"]].values,
                hovertemplate=(
                    "<b>Age %{x}</b><br>"
                    "Mean Δ: %{y:+.2f}<br>"
                    "P(Δ>0): %{customdata[2]:.0%}  |  σ: %{customdata[3]:.2f}<br>"
                    "N: %{customdata[0]:,} runs across %{customdata[1]} players<extra></extra>"
                ),
                name="Mean Δ",
            )
        )

        # Peak + decline via the shared KPI estimator (same numbers land in
        # the comparison scorecard, so the two views can never disagree).
        kpi = estimate_peak(pa)
        if not np.isnan(kpi["peak_age"]):
            fig.add_vline(
                x=kpi["peak_age"],
                line_dash="dash",
                line_color="#dc2626",
                line_width=2,
                annotation_text=f"Peak ≈ {kpi['peak_age']:.1f}",
                annotation_position="top",
                annotation_font_color="#dc2626",
            )
        if not np.isnan(kpi["decline_slope"]):
            fig.add_annotation(
                text=(
                    f"<b>Decline slope</b>: {kpi['decline_slope']:+.2f} OVR/yr<br>"
                    f"<b>R²</b> = {kpi['decline_r2']:.2f}"
                ),
                xref="paper",
                yref="paper",
                x=0.98,
                y=0.02,
                showarrow=False,
                align="right",
                font=dict(size=11, family="monospace", color="#0f172a"),
                bgcolor="rgba(255,255,255,0.92)",
                bordercolor="#e2e8f0",
                borderwidth=1,
            )

        fig.add_hline(y=0, line_width=1.5, line_color="#1e293b")
        fig.update_layout(
            template=_PLOTLY_TEMPLATE,
            title="Age Curve :: Mean OVR Δ per Year of Age (±95% CI)",
            xaxis_title="Age",
            yaxis_title="Mean OVR Delta",
        )
        return fig

    def chart_age_curve_bands(self) -> go.Figure:
        """
        Per-age P5/P50/P95 envelope + σ overlay.

        TUNING TAKEAWAY: Bands show the outcome envelope; the σ line on the
        right axis shows whether volatility is coupled to the mean (peaks
        together = single-mechanism RNG) or decoupled (different age regimes
        have independent noise). Sharp asymmetry may mean you have a special
        mechanism to suddenly boost players like godprogs!
        """
        pa = self._ds.per_age
        if pa.empty:
            return _empty_fig("Age Curve Bands")

        fig = make_subplots(specs=[[{"secondary_y": True}]])
        fig.add_trace(
            go.Scatter(
                x=pd.concat([pa["Age"], pa["Age"][::-1]]),
                y=pd.concat([pa["P95"], pa["P05"][::-1]]),
                fill="toself",
                fillcolor=hex_to_rgba("#2563eb", 0.10),
                line=dict(color="rgba(0,0,0,0)"),
                name="P5–P95",
                hoverinfo="skip",
            ),
            secondary_y=False,
        )
        for q, color, dash in [
            ("P05", "#dc2626", "dot"),
            ("P50", "#0f172a", None),
            ("P95", "#16a34a", "dot"),
        ]:
            fig.add_trace(
                go.Scatter(
                    x=pa["Age"],
                    y=pa[q],
                    mode="lines",
                    line=(
                        dict(color=color, width=2.5, dash=dash)
                        if dash
                        else dict(color=color, width=3)
                    ),
                    name=q,
                    hovertemplate=f"Age %{{x}}<br>{q}: %{{y:+.1f}}<extra></extra>",
                ),
                secondary_y=False,
            )
        fig.add_trace(
            go.Scatter(
                x=pa["Age"],
                y=pa["Std"],
                mode="lines",
                line=dict(color="#f97316", width=2, dash="dash"),
                name="σ(Δ)",
                hovertemplate="Age %{x}<br>σ: %{y:.2f}<extra></extra>",
            ),
            secondary_y=True,
        )
        fig.add_hline(y=0, line_width=1, line_color="#94a3b8", secondary_y=False)
        fig.update_xaxes(title_text="Age")
        fig.update_yaxes(title_text="OVR Δ", secondary_y=False)
        fig.update_yaxes(title_text="σ(Δ)", secondary_y=True)
        fig.update_layout(
            template=_PLOTLY_TEMPLATE,
            title="Outcome Envelope by Age (P5/Median/P95 + σ on right axis)",
            hovermode="x unified",
        )
        return fig

    def chart_age_group_curves(self) -> go.Figure:
        """Age curves per attribute group (Physical / Shooting / Mental / Skill).

        STATS NOTE: each group line is the UNWEIGHTED mean of its raw
        attribute deltas. Hence, it shows lifecycle SHAPE, not OVR impact (a big
        move on a low-coef attribute like 2Pt counts as much as a small move
        on Spd here). For OVR-impact-weighted movement, read the ΔOVR
        decomposition in §3.

        TUNING TAKEAWAY: Physicals should decline earlier and harder than
        skills. If the Mental curve looks the same as the Physical curve,
        you've collapsed two distinct lifecycle phenomena onto one schedule.
        """
        if self.ad.empty:
            return _empty_fig("Per-Group Age Curves")

        ad = self.ad.copy()
        ad["Age"] = ad["Age"].astype(int)

        groups = {
            g: [a for a in self.cols.varying_attrs if _attr_group(a) == g]
            for g in ["Physical", "Shooting", "Mental", "Skill"]
        }
        groups = {g: a for g, a in groups.items() if a}
        if not groups:
            return _empty_fig("Per-Group Age Curves")

        fig = go.Figure()
        for group, attrs in groups.items():
            color = GROUP_COLORS[group]
            # mean across the group's attributes, per age
            grouped = ad.groupby("Age")[attrs].mean().mean(axis=1)
            fig.add_trace(
                go.Scatter(
                    x=grouped.index,
                    y=grouped.values,
                    mode="lines+markers",
                    name=f"{group} ({len(attrs)} attrs)",
                    line=dict(color=color, width=3),
                    marker=dict(size=7, color=color),
                    hovertemplate=f"<b>{group}</b><br>Age %{{x}}<br>Mean Δ %{{y:+.2f}}<extra></extra>",
                )
            )

        fig.add_hline(y=0, line_width=1, line_color="#1e293b")
        fig.update_layout(
            template=_PLOTLY_TEMPLATE,
            title="Per-Group Attribute Δ by Age",
            xaxis_title="Age",
            yaxis_title="Mean Attribute Δ (averaged within group)",
            hovermode="x unified",
        )
        return fig

    # ════════════════════════════════════════════════════════════════════
    # §3  ATTRIBUTE MOVEMENT
    # ════════════════════════════════════════════════════════════════════
    def chart_attr_age_heatmap(self) -> go.Figure:
        """Age x Attribute heatmap of Mean ΔAttr.

        TUNING TAKEAWAY: Read columnwise to see where each attribute starts
        declining (the color flip); rowwise to see which attributes are net
        positive vs negative at a given age. The vertical dividers
        help you segment by group
        """
        if self.ad.empty:
            return _empty_fig("Age x Attribute Heatmap")
        avail = [c for c in self.cols.varying_attrs if c in self.ad.columns]
        if not avail:
            return _empty_fig("Age x Attribute Heatmap")

        ad = self.ad.copy()
        ad["Age"] = ad["Age"].astype(int)

        ordered = sorted(
            avail,
            key=lambda a: (
                ["Physical", "Shooting", "Mental", "Skill", "Fixed", "Other"].index(
                    _attr_group(a)
                ),
                -ad[a].abs().mean(),
            ),
        )
        pivot = ad.groupby("Age")[ordered].mean().sort_index()
        counts = ad.groupby("Age").size().reindex(pivot.index).values
        pivot = pivot.loc[counts >= Config.MIN_BIN_COUNT]
        if pivot.empty:
            return _empty_fig("Age x Attribute Heatmap")

        abs_max = max(float(np.nanpercentile(np.abs(pivot.values), 97)), 0.5)
        fig = go.Figure(
            data=go.Heatmap(
                z=pivot.values,
                x=pivot.columns,
                y=pivot.index.astype(str),
                colorscale=DIVERGING_CMAP,
                zmid=0,
                zmin=-abs_max,
                zmax=abs_max,
                text=pivot.round(2).values,
                texttemplate="%{text}",
                textfont=dict(size=9),
                hovertemplate="Age %{y} · Attr %{x}<br>Mean Δ %{z:+.2f}<extra></extra>",
                xgap=2,
                ygap=2,
                colorbar=dict(title="Mean Δ", thickness=15),
            )
        )
        groups_in_order = [_attr_group(a) for a in pivot.columns]
        prev = groups_in_order[0]
        for i, g in enumerate(groups_in_order):
            if g != prev:
                fig.add_vline(
                    x=i - 0.5, line_color="#cbd5e1", line_width=1.5, opacity=0.9
                )
                prev = g
        fig.update_layout(
            template=_PLOTLY_TEMPLATE,
            title="Age x Attribute Mean Δ (group-ordered cols)",
            xaxis_title="Attribute (light grey = group dividers)",
            yaxis_title="Age",
            height=max(450, len(pivot) * 22),
            yaxis_autorange="reversed",
            xaxis_showgrid=False,
            yaxis_showgrid=False,
        )
        return fig

    def chart_per_attribute_age_curves(self) -> go.Figure:
        """Small-multiples mean Δ vs Age, one panel per varying attribute.

        TUNING TAKEAWAY: The group chart aggregates 4 distinct curves into
        one, If Spd declines from 24 but Str only from 28, the Physical
        line averages those and hides both signals. Use this panel grid to
        spot the misbehaving attribute, then dig into the script's rule
        for it specifically.
        """
        if self.ad.empty:
            return _empty_fig("Per-Attribute Age Curves")
        avail = [c for c in self.cols.varying_attrs if c in self.ad.columns]
        if not avail:
            return _empty_fig("Per-Attribute Age Curves")

        ad = self.ad.copy()
        ad["Age"] = ad["Age"].astype(int)

        n_cols = 5
        n_rows = (len(avail) + n_cols - 1) // n_cols
        fig = make_subplots(
            rows=n_rows,
            cols=n_cols,
            subplot_titles=[f"<b>{a}</b>" for a in avail],
            vertical_spacing=0.10,
            horizontal_spacing=0.05,
        )

        for i, attr in enumerate(avail):
            r, c = (i // n_cols) + 1, (i % n_cols) + 1
            by_age = ad.groupby("Age")[attr].agg(["mean", "sem", "count"]).reset_index()
            by_age = by_age[by_age["count"] >= Config.MIN_BIN_COUNT]
            if by_age.empty:
                continue
            color = GROUP_COLORS.get(_attr_group(attr), "#475569")
            fig.add_trace(
                go.Scatter(
                    x=by_age["Age"],
                    y=by_age["mean"],
                    mode="lines+markers",
                    line=dict(color=color, width=2.5),
                    marker=dict(size=5, color=color),
                    error_y=dict(
                        type="data",
                        array=(by_age["sem"] * 1.96).fillna(0).values,
                        visible=True,
                        thickness=1,
                        width=0,
                        color=hex_to_rgba(color, 0.4),
                    ),
                    showlegend=False,
                    hovertemplate=(
                        f"<b>{attr}</b><br>Age %{{x}}<br>μΔ %{{y:+.2f}}<extra></extra>"
                    ),
                ),
                row=r,
                col=c,
            )
            fig.add_hline(y=0, line_width=0.5, line_color="#94a3b8", row=r, col=c)

        fig.update_layout(
            template=_PLOTLY_TEMPLATE,
            title="Per-Attribute Mean Δ by Age (panel color = group)",
            height=180 * n_rows,
        )
        return fig

    def chart_ovr_decomposition(self) -> go.Figure:
        """
        Per-attribute contribution to ΔOVR by age tier (coef x ΔAttr).

        TUNING TAKEAWAY: This is the only chart that tells you which attribute
        *movements* are actually shifting OVR vs which are decorative. An
        attribute can move a lot (volatility looks good) but contribute almost
        nothing to OVR because its coefficient is small. Conversely, a tiny
        movement on Hgt, oIQ, or dIQ (coef ≈ 0.15) can dominate the curve.
        Read with chart_attr_group_progression for the full picture.
        """
        if self.ad.empty:
            return _empty_fig("ΔOVR Contribution Decomposition")
        coef_map = dict(zip(Config.OVR_CALC_ORDER, Config.OVR_COEFFS))
        avail = [c for c in self.cols.varying_attrs if c in self.ad.columns]
        if not avail:
            return _empty_fig("ΔOVR Contribution Decomposition")

        ad = self.ad.copy()
        ad["AgeTier"] = Dataset.assign_tiers(ad["Age"])

        rows = []
        for tier in AGE_COLORS:
            sub = ad[ad["AgeTier"] == tier]
            if sub.empty:
                continue
            for attr in avail:
                mean_d = float(sub[attr].mean())
                coef = coef_map.get(attr, 0.0)
                rows.append(
                    {
                        "AgeTier": tier,
                        "Attribute": attr,
                        "Group": _attr_group(attr),
                        "MeanAttrDelta": mean_d,
                        "Coef": coef,
                        "OVRContribution": mean_d * coef,
                    }
                )
        data = pd.DataFrame(rows)
        if data.empty:
            return _empty_fig("ΔOVR Contribution Decomposition")

        attr_order = (
            data.groupby("Attribute")["OVRContribution"]
            .apply(lambda s: s.abs().sum())
            .sort_values(ascending=False)
            .index.tolist()
        )

        fig = go.Figure()
        for tier, color in AGE_COLORS.items():
            tdata = (
                data[data["AgeTier"] == tier]
                .set_index("Attribute")
                .reindex(attr_order)
                .reset_index()
            )
            fig.add_trace(
                go.Bar(
                    x=tdata["Attribute"],
                    y=tdata["OVRContribution"],
                    name=tier,
                    marker_color=color,
                    marker_line_width=0,
                    opacity=0.9,
                    customdata=tdata[["MeanAttrDelta", "Coef", "Group"]].values,
                    hovertemplate=(
                        "<b>%{x}</b> (%{customdata[2]})<br>"
                        "OVR Δ contribution: %{y:+.3f}<br>"
                        "Mean Δ attr: %{customdata[0]:+.2f}<br>"
                        "OVR coef: %{customdata[1]:.4f}<extra></extra>"
                    ),
                )
            )
        fig.add_hline(y=0, line_width=1, line_color="#1e293b")
        fig.update_layout(
            template=_PLOTLY_TEMPLATE,
            title="ΔOVR Contribution per Attribute (coef x ΔAttr) by Age Tier",
            xaxis_title="Attribute (sorted by total absolute OVR contribution)",
            yaxis_title="ΔOVR points contributed",
            barmode="group",
            bargap=0.2,
            bargroupgap=0.1,
        )
        return fig

    def chart_attr_comovement(self) -> go.Figure:
        """
        Two Spearman correlation matrices of attribute deltas:

        (a) BETWEEN-PLAYER -> per-player mean deltas: do players who gain on X
                            also gain on Y across the population?
        (b) WITHIN-RUN     -> avg over players of per-player run-wise corr:
                            when this player's X rolls high in a run, is Y
                            rolling high too?

        TUNING TAKEAWAY: Between-player corr exposes systematic co-movement
        from tuning (e.g., an age effect lifting multiple attrs together).
        Within-run corr exposes RNG correlation (e.g., one seed per player/run
        driving multiple attrs in lockstep). High within-run with low
        between-player = correlated noise but uncorrelated tuning.

        PERF: the within-run matrix is computed in ONE batched pass and the
        (players × runs × attrs) cube is ranked along the run axis and the
        full per-player correlation stack comes from a single einsum, instead
        of a python loop calling pandas .corr() per player. Falls back to the
        loop when players have unequal run counts or NaNs are present.
        """
        if self.ad.empty:
            return _empty_fig("Attribute Co-Movement")

        available_attributes = [
            col for col in self.cols.varying_attrs if col in self.ad.columns
        ]
        if len(available_attributes) < 2:
            return _empty_fig("Attribute Co-Movement")

        ordered_attributes = sorted(
            available_attributes,
            key=lambda attr: (
                ["Physical", "Shooting", "Mental", "Skill", "Fixed", "Other"].index(
                    _attr_group(attr)
                ),
                attr,
            ),
        )

        # 1. Between-player correlations
        between_player_corr = self.ad[ordered_attributes].corr(method="spearman").values

        # 2. Within-run correlations (averaged across players)
        sim_attributes = [
            attr for attr in ordered_attributes if attr in self.sim.columns
        ]
        n_ordered = len(ordered_attributes)
        within_run_corr = np.full((n_ordered, n_ordered), np.nan)

        if sim_attributes:
            sub = self.sim[["PlayerID", "Run"] + sim_attributes]
            counts = sub.groupby("PlayerID").size()
            equal_runs = counts.nunique() == 1 and counts.iloc[0] >= 5
            has_nan = sub[sim_attributes].isna().any().any()

            if equal_runs and not has_nan:
                n_runs = int(counts.iloc[0])
                arr = sub.sort_values(["PlayerID", "Run"])[
                    sim_attributes
                ].values.reshape(-1, n_runs, len(sim_attributes))
                ranks = scipy_stats.rankdata(arr, axis=1)
                Z = ranks - ranks.mean(axis=1, keepdims=True)
                num = np.einsum("pri,prj->pij", Z, Z, optimize=True)
                ss = (Z**2).sum(axis=1)  # (P, A)
                denom = np.sqrt(ss[:, :, None] * ss[:, None, :])
                with np.errstate(invalid="ignore", divide="ignore"):
                    corr = num / denom
                corr[~np.isfinite(corr)] = np.nan
                mapped = np.nanmean(corr, axis=0)  # (A, A)
            else:
                # Fallback: original per-player loop (rare path)
                acc = np.zeros((len(sim_attributes),) * 2)
                cnt = np.zeros_like(acc)
                for _, g in sub.groupby("PlayerID"):
                    clean = g[sim_attributes].dropna()
                    if len(clean) < 5:
                        continue
                    m = clean.corr(method="spearman").values
                    mask = ~np.isnan(m)
                    acc[mask] += m[mask]
                    cnt[mask] += 1
                mapped = np.where(cnt > 0, acc / np.maximum(cnt, 1), np.nan)

            idx = [ordered_attributes.index(a) for a in sim_attributes]
            within_run_corr[np.ix_(idx, idx)] = mapped

        # 3. Plot
        fig = make_subplots(
            rows=1,
            cols=2,
            subplot_titles=(
                "<b>(a) Between-player</b>",
                "<b>(b) Within-run (avg over players)</b>",
            ),
            horizontal_spacing=0.08,
        )
        for idx2, correlation_matrix in enumerate(
            [between_player_corr, within_run_corr]
        ):
            is_second_subplot = idx2 == 1
            fig.add_trace(
                go.Heatmap(
                    z=correlation_matrix,
                    x=ordered_attributes,
                    y=ordered_attributes,
                    colorscale=DIVERGING_CMAP,
                    zmid=0,
                    zmin=-1,
                    zmax=1,
                    hovertemplate="%{y} x %{x}<br>ρ = %{z:+.3f}<extra></extra>",
                    xgap=3,
                    ygap=3,
                    text=np.round(correlation_matrix, 2),
                    texttemplate="%{text}",
                    textfont=dict(size=9),
                    showscale=is_second_subplot,
                    colorbar=(
                        dict(title="Spearman ρ", thickness=15)
                        if is_second_subplot
                        else None
                    ),
                ),
                row=1,
                col=idx2 + 1,
            )
            fig.update_yaxes(autorange="reversed", row=1, col=idx2 + 1)

        fig.update_layout(
            template=_PLOTLY_TEMPLATE,
            title="Attribute Co-Movement :: Between-Player vs Within-Run",
            height=620,
        )
        return fig

    def chart_attribute_saturation(self) -> go.Figure:
        """
        % of (player, run) attribute values pinned at boundaries.

        TUNING TAKEAWAY: A clamp that fires too often is invisible everywhere
        else in this dashboard. If 40% of young players are ceiling-pinned on
        3Pt at 99, the script literally cannot move them. This chart surfaces that.
        Floor saturation on physicals for the Oldest tier is expected; floor
        saturation on Mental attrs is usually a bug I would imagine.
        """
        avail = [c for c in self.cols.varying_attrs if c in self.sim.columns]
        if not avail:
            return _empty_fig("Attribute Boundary Saturation")

        df = self.sim
        rows = []
        for tier in AGE_COLORS:
            sub = df[df["AgeTier"] == tier]
            if sub.empty:
                continue
            for attr in avail:
                vals = sub[attr].dropna().values
                if len(vals) == 0:
                    continue
                rows.append(
                    {
                        "AgeTier": tier,
                        "Attribute": attr,
                        "PctCeiling": float((vals >= 99).mean()),
                        "PctFloor": float((vals <= 1).mean()),
                        "N": len(vals),
                    }
                )
        data = pd.DataFrame(rows)
        if data.empty:
            return _empty_fig("Attribute Boundary Saturation")

        fig = make_subplots(
            rows=1,
            cols=len(AGE_COLORS),
            subplot_titles=[f"<b>{t}</b>" for t in AGE_COLORS],
            horizontal_spacing=0.04,
            shared_yaxes=True,
        )
        x_max = max(0.05, data["PctCeiling"].max() * 1.1, data["PctFloor"].max() * 1.1)

        for i, tier in enumerate(AGE_COLORS):
            sub = data[data["AgeTier"] == tier].sort_values("Attribute")
            if sub.empty:
                continue
            fig.add_trace(
                go.Bar(
                    y=sub["Attribute"],
                    x=sub["PctCeiling"],
                    orientation="h",
                    name="At ceiling (≥99)",
                    marker_color="#2563eb",
                    opacity=0.9,
                    showlegend=(i == 0),
                    customdata=sub["N"].values,
                    hovertemplate="%{y} :: ceiling %{x:.1%} (N=%{customdata:,})<extra></extra>",
                ),
                row=1,
                col=i + 1,
            )
            fig.add_trace(
                go.Bar(
                    y=sub["Attribute"],
                    x=-sub["PctFloor"],
                    orientation="h",
                    name="At floor (≤1)",
                    marker_color="#dc2626",
                    opacity=0.9,
                    showlegend=(i == 0),
                    customdata=sub[["PctFloor", "N"]].values,
                    hovertemplate="%{y} :: floor %{customdata[0]:.1%} (N=%{customdata[1]:,})<extra></extra>",
                ),
                row=1,
                col=i + 1,
            )
            fig.update_xaxes(row=1, col=i + 1, tickformat=".0%", range=[-x_max, x_max])

        fig.update_layout(
            template=_PLOTLY_TEMPLATE,
            title="Attribute Boundary Saturation :: % Floor (<--) vs % Ceiling (-->)",
            barmode="overlay",
            height=max(450, len(avail) * 22),
        )
        return fig

    def chart_ovr_variance_decomposition(self) -> go.Figure:
        """
        Per-attribute contribution to Var(ΔOVR) via c_i · Cov(Δ_i, ΔOVR).

        TUNING TAKEAWAY: chart_ovr_decomposition tells you which attributes
        drive the *level* of ΔOVR. This one tells you which attributes
        drive the SPREAD of ΔOVR ACROSS PLAYERS, i.e. why some players' mean
        movement is bigger than others'. By the linearity of covariance the
        bars sum exactly to Var(ΔOVR_implied): Var(Σ cᵢΔᵢ) = Σ cᵢ·Cov(Δᵢ,ΔOVR),
        a clean 100% attribution. A single dominant bar = between-player
        dispersion carried by one attribute. Negative bars = that attribute
        moves opposite to ΔOVR across players.

        STATS NOTE: this is BETWEEN-PLAYER dispersion of each player's MEAN
        movement, not run-to-run RNG uncertainty. Run-to-run noise is the
        within-player variance shown by the ICC chart.
        """
        if self.ad.empty:
            return _empty_fig("OVR Variance Decomposition")
        coef_map = dict(zip(Config.OVR_CALC_ORDER, Config.OVR_COEFFS))
        avail = [c for c in self.cols.varying_attrs if c in self.ad.columns]
        if not avail:
            return _empty_fig("OVR Variance Decomposition")

        coefs = np.array([coef_map[a] for a in avail])
        attr_mat = self.ad[avail].dropna().values
        if len(attr_mat) < 3:
            return _empty_fig("OVR Variance Decomposition")
        implied = attr_mat @ coefs
        total_var = float(implied.var(ddof=1))
        rows = []
        for i, a in enumerate(avail):
            cov_i = float(np.cov(attr_mat[:, i], implied, ddof=1)[0, 1])
            rows.append(
                {
                    "Attribute": a,
                    "Group": _attr_group(a),
                    "Contribution": coefs[i] * cov_i,
                    "OwnVar": coefs[i] ** 2 * float(np.var(attr_mat[:, i], ddof=1)),
                    "Coef": float(coefs[i]),
                }
            )
        data = pd.DataFrame(rows).sort_values("Contribution", ascending=True)
        data["Pct"] = 100 * data["Contribution"] / total_var if total_var > 0 else 0
        colors = [GROUP_COLORS.get(g, "#475569") for g in data["Group"]]

        fig = go.Figure(
            go.Bar(
                y=data["Attribute"],
                x=data["Contribution"],
                orientation="h",
                marker_color=colors,
                marker_line_width=0,
                opacity=0.9,
                text=[f"{p:+.0f}%" for p in data["Pct"]],
                textposition="outside",
                textfont=dict(size=10),
                customdata=data[["Group", "Pct", "OwnVar", "Coef"]].values,
                hovertemplate=(
                    "<b>%{y}</b> (%{customdata[0]})<br>"
                    "Total contribution: %{x:+.4f}  "
                    "(%{customdata[1]:+.0f}%)<br>"
                    "Own-variance only: %{customdata[2]:.4f}<br>"
                    "Coef: %{customdata[3]:.4f}<extra></extra>"
                ),
            )
        )
        fig.add_vline(x=0, line_width=1, line_color="#1e293b")
        top = data.reindex(data["Contribution"].abs().sort_values().index).iloc[-1]
        fig.add_annotation(
            text=(
                f"<b>Var(ΔOVR)</b> = {total_var:.3f} across players<br>"
                f"<b>{top['Attribute']}</b> carries {top['Pct']:+.0f}% of the spread"
            ),
            xref="paper",
            yref="paper",
            x=0.98,
            y=0.02,
            showarrow=False,
            align="right",
            font=dict(size=11, family="monospace", color="#0f172a"),
            bgcolor="rgba(255,255,255,0.92)",
            bordercolor="#e2e8f0",
            borderwidth=1,
        )
        fig.update_layout(
            template=_PLOTLY_TEMPLATE,
            title="ΔOVR Variance Decomposition (bars sum to Var(ΔOVR))",
            xaxis_title="Contribution :: c_i x Cov(ΔAttr_i, ΔOVR)",
            yaxis_title="",
            height=max(450, len(data) * 28),
        )
        return fig

    # ════════════════════════════════════════════════════════════════════
    # §4  PER-PLAYER OUTCOMES
    # ════════════════════════════════════════════════════════════════════

    def chart_risk_return(self) -> go.Figure:
        """
        μ(Δ) vs σ(Δ) per player, coloured by baseline OVR.

        TUNING TAKEAWAY: the top-right quadrant (high gain, high risk)
        and the bottom-right quadrant (high gain, low risk)
        with most of the young high-baseline cohort.
        """
        ppl = self.ppl.dropna(subset=["MeanDelta", "StdDelta"])
        if len(ppl) < 5:
            return _empty_fig("Risk-Return")

        fig = go.Figure()
        fig.add_trace(
            go.Scatter(
                x=ppl["MeanDelta"],
                y=ppl["StdDelta"],
                mode="markers",
                marker=dict(
                    color=ppl["Baseline"],
                    colorscale="RdYlBu_r",
                    size=10,
                    line=dict(width=1, color="white"),
                    opacity=0.85,
                    colorbar=dict(title="Baseline OVR", thickness=15, len=0.7),
                    showscale=True,
                ),
                customdata=ppl[
                    ["Label", "Age", "Baseline", "PctPositive", "AgeTier"]
                ].values,
                hovertemplate=(
                    "<b>%{customdata[0]}</b><br>"
                    "Age %{customdata[1]}  |  Base %{customdata[2]:.0f}  |  %{customdata[4]}<br>"
                    "Mean Δ %{x:+.1f}  |  σ %{y:.1f}<br>"
                    "P(improve): %{customdata[3]:.0%}<extra></extra>"
                ),
            )
        )

        med_x = ppl["MeanDelta"].median()
        med_y = ppl["StdDelta"].median()
        fig.add_hline(
            y=med_y,
            line_dash="dash",
            line_color="#94a3b8",
            annotation_text=f"Median σ: {med_y:.1f}",
        )
        fig.add_vline(
            x=med_x,
            line_dash="dash",
            line_color="#94a3b8",
            annotation_text=f"Median μ: {med_x:.1f}",
        )
        fig.add_vline(x=0, line_width=2, line_color="#1e293b")

        for xq, yq, txt in [
            (0.85, 0.15, "★ STAR  (high gain · low risk)"),
            (0.85, 0.85, "⚡ VOLATILE  (high gain · high risk)"),
            (0.15, 0.15, "■ STABLE  (low gain · low risk)"),
            (0.15, 0.85, "▽ DANGER  (low gain · high risk)"),
        ]:
            fig.add_annotation(
                x=ppl["MeanDelta"].quantile(xq),
                y=ppl["StdDelta"].quantile(yq),
                text=txt,
                showarrow=False,
                font=dict(size=10, color="#475569"),
                bgcolor="rgba(255,255,255,0.75)",
            )

        fig.update_layout(
            template=_PLOTLY_TEMPLATE,
            title="Risk-Return Map (μΔ vs σΔ, coloured by Baseline OVR)",
            xaxis_title="Mean OVR Δ",
            yaxis_title="Std of OVR Δ",
        )
        return fig

    def chart_outcome_distributions(self) -> go.Figure:
        """KDE of every (player, run) Δ by age tier.

        TUNING TAKEAWAY: Look at the relative position of the tier means
        and the asymmetry of the tails. A long right tail in the Youngest
        tier is the god-prog signal; a long left tail in the Oldest tier
        is the catastrophic decline signal.
        """
        df = self.sim
        fig = go.Figure()
        tier_groups: Dict[str, np.ndarray] = {}
        for tier, color in AGE_COLORS.items():
            vals = df[df["AgeTier"] == tier]["Delta"].dropna().values
            if len(vals) < Config.MIN_KDE_SAMPLES:
                continue
            tier_groups[tier] = vals
            if np.std(vals) < 1e-10:
                fig.add_vline(
                    x=vals[0],
                    line_dash="dash",
                    line_color=color,
                    annotation_text=f"{tier}: const {vals[0]:.1f}",
                )
                continue
            kde = gaussian_kde(vals, bw_method="scott")
            xgrid = np.linspace(vals.min() - 3, vals.max() + 3, 320)
            fig.add_trace(
                go.Scatter(
                    x=xgrid,
                    y=kde(xgrid),
                    mode="lines",
                    name=tier,
                    fill="tozeroy",
                    fillcolor=hex_to_rgba(color, 0.15),
                    line=dict(color=color, width=3, shape="spline"),
                    hovertemplate=f"<b>{tier}</b><br>Δ %{{x:+.1f}}<br>density %{{y:.4f}}<extra></extra>",
                )
            )
            fig.add_vline(
                x=float(vals.mean()),
                line_dash="dash",
                line_color=color,
                line_width=1.5,
                annotation_text=f"μ={vals.mean():+.1f}",
                annotation_position="top",
                annotation_font_color=color,
            )

        # Pairwise KS tests
        names = list(tier_groups.keys())
        rows = []
        for i in range(len(names)):
            for j in range(i + 1, len(names)):
                ks, p = scipy_stats.ks_2samp(
                    tier_groups[names[i]], tier_groups[names[j]]
                )
                d = cohens_d(tier_groups[names[i]], tier_groups[names[j]])
                rows.append(
                    f"{names[i]} vs {names[j]}: KS={ks:.3f} (p={p:.4f}), d={d:+.2f}"
                )
        if rows:
            fig.add_annotation(
                text="<br>".join(rows),
                xref="paper",
                yref="paper",
                x=0.98,
                y=0.98,
                showarrow=False,
                align="right",
                font=dict(size=10, family="monospace", color="#64748b"),
                bgcolor="rgba(255,255,255,0.92)",
                bordercolor="#e2e8f0",
                borderwidth=1,
            )
        fig.add_vline(x=0, line_width=2, line_color="#1e293b")
        fig.update_layout(
            template=_PLOTLY_TEMPLATE,
            title="Outcome Distributions by Age Tier",
            xaxis_title="OVR Δ",
            yaxis_title="Probability Density",
            hovermode="x unified",
        )
        return fig

    def chart_improve_probability(self) -> go.Figure:
        """P(Δ > 0) as a function of baseline OVR, with logistic fit per tier.

        TUNING TAKEAWAY: A clean monotonic decline from low baseline to
        high baseline says "the cap bites". Tier-specific deviations (e.g.
        Youngest staying near 100% across baselines) reveal that the
        age effect dominates the cap for that bracket.
        """
        ppl = self.ppl.dropna(subset=["Baseline", "PctPositive"])
        if len(ppl) < 15:
            return _empty_fig("Improvement Probability")

        fig = go.Figure()
        for tier, color in AGE_COLORS.items():
            sub = ppl[ppl["AgeTier"] == tier]
            if len(sub) < 5:
                continue
            fig.add_trace(
                go.Scatter(
                    x=sub["Baseline"],
                    y=sub["PctPositive"],
                    mode="markers",
                    marker=dict(
                        color=color,
                        size=8,
                        opacity=0.65,
                        line=dict(width=1, color="white"),
                    ),
                    name=tier,
                    customdata=sub[["Label", "Age"]].values,
                    hovertemplate=(
                        "<b>%{customdata[0]}</b><br>Age %{customdata[1]}<br>"
                        "Base %{x:.0f}  |  P(+) %{y:.0%}<extra></extra>"
                    ),
                )
            )
            if len(sub) >= 10:
                try:
                    popt, _ = curve_fit(
                        logistic,
                        sub["Baseline"],
                        sub["PctPositive"],
                        p0=[0, -0.1],
                        maxfev=4000,
                    )
                    xfit = np.linspace(
                        sub["Baseline"].min(), sub["Baseline"].max(), 200
                    )
                    fig.add_trace(
                        go.Scatter(
                            x=xfit,
                            y=logistic(xfit, *popt),
                            mode="lines",
                            line=dict(color=color, width=2.5, dash="dash"),
                            name=f"{tier} fit",
                            showlegend=False,
                            hoverinfo="skip",
                        )
                    )
                except RuntimeError:
                    pass

        fig.add_hline(
            y=0.5,
            line_width=1,
            line_dash="dot",
            line_color="#94a3b8",
            annotation_text="50% threshold",
        )
        fig.update_layout(
            template=_PLOTLY_TEMPLATE,
            title="Probability of Improvement by Baseline OVR",
            xaxis_title="Baseline OVR",
            yaxis_title="P(Δ > 0)",
            yaxis_tickformat=".0%",
        )
        return fig

    # ════════════════════════════════════════════════════════════════════
    # §5  INPUT SENSITIVITY
    # ════════════════════════════════════════════════════════════════════
    def chart_controlled_input_effects(self) -> go.Figure:
        """Standardized OLS coefficient for each input on Mean Δ,
        MARGINALLY controlling for Age + Baseline, with bootstrap 95% CIs.

        TUNING TAKEAWAY: Shows which inputs track outcomes beyond age and
        baseline. An input whose CI straddles zero has no detectable
        marginal effect.

        STATS NOTE: each input is fit in its OWN model (input + Age +
        Baseline), so this is a MARGINAL effect, not a unique one. BBGM's
        advanced stats are heavily collinear (EWA is a pure function of PER;
        DWS tracks DBPM), so a joint model has VIFs in the millions and
        unstable betas. We instead collapse near-duplicate inputs (|r|≥0.95)
        into one representative per cluster and note the members. For a
        credit-splitting importance that IS collinearity-robust, read the
        Shapley R² chart next to this one.
        """
        player_data = self.ppl

        # Filter for varying configuration inputs present in the dataset
        all_valid = [
            col
            for col in self.cols.inputs
            if col in player_data.columns and player_data[col].std() > 0
        ]
        if not all_valid or len(player_data) < 20:
            return _empty_fig("Controlled Input Effects")

        # Collapse near-duplicate inputs: greedy clustering on |Pearson r|.
        # Keeps the first-seen member as the representative; the rest are
        # recorded so the annotation can say "PER (≡ EWA)".
        corr = player_data[all_valid].corr().abs()
        assigned: set = set()
        clusters: List[List[str]] = []
        for col in all_valid:
            if col in assigned:
                continue
            members = [col] + [
                o
                for o in all_valid
                if o != col and o not in assigned and corr.loc[col, o] >= 0.95
            ]
            assigned.update(members)
            clusters.append(members)
        valid_inputs = [c[0] for c in clusters]
        collapsed = {c[0]: c[1:] for c in clusters if len(c) > 1}

        # Set up basic baseline control features
        control_features = ["Age"]
        if "Baseline" in player_data.columns and player_data["Baseline"].std() > 0:
            control_features.append("Baseline")

        # Standardize the target variable (Mean Delta)
        standardized_target = zscore(player_data["MeanDelta"]).values

        sample_size = len(player_data)
        random_generator = np.random.default_rng(42)
        n_bootstrap = 600

        regression_results = []

        for input_feature in valid_inputs:
            model_features = control_features + [input_feature]
            standardized_columns = [
                zscore(player_data[feat]).values for feat in model_features
            ]

            # Build design matrix with an intercept column
            design_matrix = np.column_stack(
                [np.ones(sample_size)] + standardized_columns
            )

            try:
                coefficients, _, r_squared = ols_fit(design_matrix, standardized_target)
            except (ValueError, np.linalg.LinAlgError):
                continue

            bootstrap_coefficients = []
            for _ in range(n_bootstrap):
                resample_indices = random_generator.integers(
                    0, sample_size, sample_size
                )
                try:
                    # Fit model on the bootstrapped resample selection
                    boot_coefficients, _, _ = ols_fit(
                        design_matrix[resample_indices],
                        standardized_target[resample_indices],
                        compute_se=False,
                    )
                    bootstrap_coefficients.append(boot_coefficients[-1])
                except (ValueError, np.linalg.LinAlgError):
                    continue

            if len(bootstrap_coefficients) < 50:
                continue

            lower_bound, upper_bound = np.percentile(
                bootstrap_coefficients, [2.5, 97.5]
            )
            regression_results.append(
                {
                    "Input": input_feature,
                    "Beta": float(coefficients[-1]),
                    "CI_lower": float(lower_bound),
                    "CI_upper": float(upper_bound),
                    "R2": r_squared,
                }
            )

        if not regression_results:
            return _empty_fig("Controlled Input Effects")

        results_df = pd.DataFrame(regression_results)
        # Annotate representatives that stand in for a collinear cluster.
        results_df["Label"] = results_df["Input"].apply(
            lambda c: f"{c}  (≡ {', '.join(collapsed[c])})" if c in collapsed else c
        )
        # Sort features by the magnitude of their standardized effect size (absolute beta)
        results_df = results_df.reindex(results_df["Beta"].abs().sort_values().index)

        fig = go.Figure()

        # Visual color design configuration
        COLOR_POSITIVE_SIGNIFICANT = "#16a34a"  # Green
        COLOR_NEGATIVE_SIGNIFICANT = "#dc2626"  # Red
        COLOR_INSIGNIFICANT = "#94a3b8"  # Slate gray

        for _, row in results_df.iterrows():
            # Check if the confidence interval entirely excludes/straddles zero
            is_statistically_significant = (row["CI_lower"] > 0) or (
                row["CI_upper"] < 0
            )

            if is_statistically_significant:
                line_color = (
                    COLOR_POSITIVE_SIGNIFICANT
                    if row["Beta"] > 0
                    else COLOR_NEGATIVE_SIGNIFICANT
                )
            else:
                line_color = COLOR_INSIGNIFICANT

            # Draw error bar horizontal line segment
            fig.add_trace(
                go.Scatter(
                    x=[row["CI_lower"], row["CI_upper"]],
                    y=[row["Label"], row["Label"]],
                    mode="lines",
                    line=dict(color=line_color, width=3),
                    showlegend=False,
                    hoverinfo="skip",
                )
            )

            # Draw target central point marker
            fig.add_trace(
                go.Scatter(
                    x=[row["Beta"]],
                    y=[row["Label"]],
                    mode="markers",
                    marker=dict(
                        color=line_color, size=12, line=dict(width=2, color="white")
                    ),
                    customdata=[[row["CI_lower"], row["CI_upper"], row["R2"]]],
                    hovertemplate=(
                        "<b>%{y}</b><br>"
                        "Std β = %{x:+.3f}<br>"
                        "95% CI = [%{customdata[0]:+.3f}, %{customdata[1]:+.3f}]<br>"
                        "Full-model R² = %{customdata[2]:.3f}<extra></extra>"
                    ),
                    showlegend=False,
                )
            )

        # Vertical anchor line at 0 indicating null hypothesis boundary
        fig.add_vline(x=0, line_width=1.5, line_color="#1e293b")

        if collapsed:
            note = "<br>".join(f"{k} ≡ {', '.join(v)}" for k, v in collapsed.items())
            fig.add_annotation(
                text=(f"<b>Collinear inputs merged (|r|≥0.95):</b><br>{note}"),
                xref="paper",
                yref="paper",
                x=0.02,
                y=0.02,
                showarrow=False,
                align="left",
                font=dict(size=10, family="monospace", color="#64748b"),
                bgcolor="rgba(255,255,255,0.92)",
                bordercolor="#e2e8f0",
                borderwidth=1,
            )

        fig.update_layout(
            template=_PLOTLY_TEMPLATE,
            title=(
                f"Marginal Effect of Each Input on Mean Δ "
                f"(each controlling for {', '.join(control_features)})"
            ),
            xaxis_title="Standardized β (effect on Mean Δ per 1-SD of input)",
            yaxis_title="",
            height=max(400, len(results_df) * 34),
        )
        return fig

    def chart_incremental_r2(self) -> go.Figure:
        """Shapley-averaged R² importance over Mean Δ.

        TUNING TAKEAWAY: Each bar = the average marginal R² that input
        contributes, averaged over many random inclusion orderings. Shapley
        splits credit symmetrically. Bars here are STABLE importance,
        where the order on the y-axis reflects magnitude, not entry sequence.

        PERF: subset-R² values are cached, so the cost is bounded by the
        number of distinct feature subsets. Candidates are capped at the 12
        strongest |spearman| correlates so the subset space (2^k) can't
        explode as more model inputs are added upstream.
        """
        ppl = self.ppl
        candidates = [
            c for c in self.cols.all_predictors if c in ppl.columns and ppl[c].std() > 0
        ]
        if len(candidates) < 2:
            return _empty_fig("Shapley R²")

        y = ppl["MeanDelta"].values
        if len(candidates) > 12:
            strength = {
                c: abs(scipy_stats.spearmanr(ppl[c].values, y)[0]) for c in candidates
            }
            candidates = sorted(candidates, key=lambda c: -np.nan_to_num(strength[c]))[
                :12
            ]

        n = len(ppl)
        cached: Dict[frozenset, float] = {frozenset(): 0.0}

        def r2_for(features: List[str]) -> float:
            key = frozenset(features)
            if key in cached:
                return cached[key]
            X = np.column_stack(
                [np.ones(n)] + [zscore(ppl[f]).values for f in features]
            )
            try:
                _, _, r2 = ols_fit(X, y, compute_se=False)
                r2 = max(0.0, r2)
            except (ValueError, np.linalg.LinAlgError):
                r2 = 0.0
            cached[key] = r2
            return r2

        rng = np.random.default_rng(0)
        n_orderings = 200
        contribs = {c: [] for c in candidates}
        for _ in range(n_orderings):
            order = list(candidates)
            rng.shuffle(order)
            included: List[str] = []
            prev = 0.0
            for feat in order:
                included.append(feat)
                curr = r2_for(included)
                contribs[feat].append(curr - prev)
                prev = curr

        data = pd.DataFrame(
            {
                "Input": list(contribs.keys()),
                "Shapley": [float(np.mean(v)) for v in contribs.values()],
                "SE": [float(np.std(v) / np.sqrt(len(v))) for v in contribs.values()],
            }
        ).sort_values("Shapley", ascending=True)
        total_r2 = float(data["Shapley"].sum())
        colors = ["#2563eb" if v > 0.01 else "#94a3b8" for v in data["Shapley"]]

        fig = go.Figure(
            go.Bar(
                y=data["Input"],
                x=data["Shapley"],
                orientation="h",
                marker_color=colors,
                marker_line_width=0,
                opacity=0.9,
                error_x=dict(
                    type="data",
                    array=(data["SE"] * 1.96).values,
                    visible=True,
                    thickness=1.5,
                    width=3,
                ),
                text=[f"{v:.3f}" for v in data["Shapley"]],
                textposition="outside",
                textfont=dict(size=10),
                hovertemplate="<b>%{y}</b><br>Shapley R²: %{x:.4f}<extra></extra>",
            )
        )
        fig.add_annotation(
            text=f"<b>Total R²</b>: {total_r2:.3f}<br><b>Orderings</b>: {n_orderings}",
            xref="paper",
            yref="paper",
            x=0.98,
            y=0.02,
            showarrow=False,
            align="right",
            font=dict(size=11, family="monospace", color="#0f172a"),
            bgcolor="rgba(255,255,255,0.92)",
            bordercolor="#e2e8f0",
            borderwidth=1,
        )
        fig.update_layout(
            template=_PLOTLY_TEMPLATE,
            title="Shapley-Averaged R² Importance over Mean Δ",
            xaxis_title="Avg marginal R² (stable across orderings)",
            yaxis_title="",
            showlegend=False,
            height=max(400, len(data) * 30),
        )
        return fig

    def chart_partial_dependence(self) -> go.Figure:
        """Partial dependence on the Age- and Baseline-residualized Mean Δ.

        TUNING TAKEAWAY: After regressing out Age + Baseline, the residual
        in Mean Δ is what the input is being asked to explain. A clean
        linear ascent = the input is wired linearly INDEPENDENT of age and
        baseline. A flat region with a kink = the script has a threshold.
        A bump = non-monotonic.
        """
        ppl = self.ppl
        candidates = [
            c for c in self.cols.inputs if c in ppl.columns and ppl[c].std() > 0
        ]
        if not candidates:
            return _empty_fig("Partial Dependence")

        controls = ["Age"]
        if "Baseline" in ppl.columns and ppl["Baseline"].std() > 0:
            controls.append("Baseline")
        X_ctrl = np.column_stack(
            [np.ones(len(ppl))] + [zscore(ppl[c]).values for c in controls]
        )
        y = ppl["MeanDelta"].values
        try:
            beta, _, _ = ols_fit(X_ctrl, y)
            y_resid = y - X_ctrl @ beta
        except (ValueError, np.linalg.LinAlgError):
            y_resid = y - y.mean()
        ppl_r = ppl.copy()
        ppl_r["_resid"] = y_resid

        scored = [
            (c, abs(scipy_stats.spearmanr(ppl_r[c], ppl_r["_resid"])[0]))
            for c in candidates
        ]
        scored = [s for s in scored if not np.isnan(s[1])]
        scored.sort(key=lambda x: x[1], reverse=True)
        top = [s[0] for s in scored[:6]]
        if not top:
            return _empty_fig("Partial Dependence")

        cols = min(3, len(top))
        rows = (len(top) + cols - 1) // cols
        fig = make_subplots(
            rows=rows,
            cols=cols,
            subplot_titles=[f"<b>{d}</b>" for d in top],
            vertical_spacing=0.14,
            horizontal_spacing=0.08,
        )

        for i, driver in enumerate(top):
            r, c = (i // cols) + 1, (i % cols) + 1
            local = ppl_r.dropna(subset=[driver, "_resid"]).copy()
            q = qcut_safe(local[driver], 8)
            if q is None:
                continue
            local["Bin"] = q
            binned = (
                local.groupby("Bin", observed=True)
                .agg(
                    Xmid=(driver, "mean"),
                    Y=("_resid", "mean"),
                    Yse=("_resid", "sem"),
                    N=("_resid", "count"),
                )
                .reset_index()
            )
            fig.add_trace(
                go.Scatter(
                    x=binned["Xmid"],
                    y=binned["Y"],
                    mode="lines+markers",
                    line=dict(color="#0f172a", width=3),
                    marker=dict(size=8, color="#0f172a"),
                    error_y=dict(
                        type="data",
                        array=(binned["Yse"] * 1.96).fillna(0).values,
                        visible=True,
                        thickness=1.5,
                        width=3,
                    ),
                    showlegend=False,
                    hovertemplate=(
                        f"<b>{driver}</b><br>bin midpoint %{{x:.2f}}<br>"
                        f"residual μΔ %{{y:+.2f}}<br>N=%{{customdata}}<extra></extra>"
                    ),
                    customdata=binned["N"].values,
                ),
                row=r,
                col=c,
            )
            fig.add_hline(y=0, line_width=0.5, line_color="#94a3b8", row=r, col=c)

        fig.update_layout(
            template=_PLOTLY_TEMPLATE,
            title=(
                f"Partial Dependence :: Residualized Mean Δ vs Input Quantile "
                f"(controls: {', '.join(controls)})"
            ),
            height=320 * rows,
        )
        return fig

    # ════════════════════════════════════════════════════════════════════
    # §6  RNG CALIBRATION
    # ════════════════════════════════════════════════════════════════════

    def chart_icc(self) -> go.Figure:
        """ICC per varying attribute: σ²_between / total variance.

        TUNING TAKEAWAY: Aim for the bulk of bars in the green band
        (0.4-0.85). Below 0.4 means the script is so random that player
        identity is barely affecting outcomes. Above 0.9 means the
        script is effectively deterministic.
        """
        icc = self._ds.icc
        if icc.empty:
            return _empty_fig("ICC")

        fig = go.Figure()
        colors = [
            "#16a34a" if v >= 0.7 else "#f97316" if v >= 0.4 else "#dc2626"
            for v in icc["ICC"]
        ]
        fig.add_trace(
            go.Bar(
                y=icc["Attribute"],
                x=icc["ICC"],
                orientation="h",
                marker_color=colors,
                marker_line_width=0,
                opacity=0.9,
                customdata=icc[["Var_Between", "Var_Within", "Group"]].values,
                hovertemplate=(
                    "<b>%{y}</b> (%{customdata[2]})<br>"
                    "ICC %{x:.2f}<br>"
                    "σ²_between %{customdata[0]:.2f}<br>"
                    "σ²_within  %{customdata[1]:.2f}<extra></extra>"
                ),
            )
        )
        fig.add_vline(
            x=0.4,
            line_dash="dot",
            line_color="#f97316",
            annotation_text="Low (0.4)",
            annotation_position="top right",
        )
        fig.add_vline(
            x=0.7,
            line_dash="dot",
            line_color="#16a34a",
            annotation_text="High (0.7)",
            annotation_position="top right",
        )

        fig.update_layout(
            template=_PLOTLY_TEMPLATE,
            title="Variance Decomposition :: ICC per Attribute",
            xaxis_title="ICC (0 = pure RNG, 1 = fully deterministic)",
            yaxis_title="",
            height=max(400, len(icc) * 26),
        )
        return fig

    def chart_convergence(self) -> go.Figure:
        """Running mean ± MCSE for the most volatile players.

        TUNING TAKEAWAY: If the bands haven't visibly narrowed by the
        end of the run, you haven't done enough Monte Carlo passes. The
        target line (red dashed) shows where the running mean settled.
        """
        conv = self._ds.convergence
        pids = conv["top_volatile"]
        if not pids:
            return _empty_fig("Convergence")

        labels = self.ppl.set_index("PlayerID")["Label"]

        fig = make_subplots(
            rows=2,
            cols=3,
            subplot_titles=[f"<b>{labels.get(p, p)}</b>" for p in pids],
            vertical_spacing=0.15,
            horizontal_spacing=0.1,
        )
        for idx, pid in enumerate(pids):
            r, c = (idx // 3) + 1, (idx % 3) + 1
            v = self.sim[self.sim["PlayerID"] == pid].sort_values("Run")["Delta"].values
            n = len(v)
            if n < 2:
                continue
            runs = np.arange(1, n + 1)
            cum_sum = np.cumsum(v)
            cum_sum_sq = np.cumsum(v**2)
            cum_mean = cum_sum / runs
            var_pop = cum_sum_sq / runs - cum_mean**2
            safe = np.where(runs > 1, runs - 1, 1.0)
            cum_var = np.where(runs > 1, var_pop * runs / safe, 0.0)
            cum_var = np.maximum(cum_var, 0.0)
            mcse = np.sqrt(cum_var / runs)
            fig.add_trace(
                go.Scatter(
                    x=runs,
                    y=cum_mean,
                    mode="lines",
                    line=dict(color="#2563eb", width=2),
                    showlegend=False,
                ),
                row=r,
                col=c,
            )
            fig.add_trace(
                go.Scatter(
                    x=runs,
                    y=cum_mean - mcse,
                    mode="lines",
                    line=dict(width=0),
                    showlegend=False,
                ),
                row=r,
                col=c,
            )
            fig.add_trace(
                go.Scatter(
                    x=runs,
                    y=cum_mean + mcse,
                    mode="lines",
                    line=dict(width=0),
                    fill="tonexty",
                    fillcolor="rgba(37,99,235,0.18)",
                    showlegend=False,
                ),
                row=r,
                col=c,
            )
            fig.add_hline(
                y=float(np.mean(v)),
                line_dash="dash",
                line_color="#dc2626",
                line_width=1,
                row=r,
                col=c,
            )

        pct = conv["pct_converged"]
        fig.add_annotation(
            text=(
                f"<b>Convergence</b>: {pct:.0%} of players with "
                f"MCSE &lt; {Config.MCSE_THRESHOLD}  ·  max MCSE = {conv['max_mcse']:.2f}"
            ),
            xref="paper",
            yref="paper",
            x=0.5,
            y=-0.05,
            showarrow=False,
            font=dict(size=12, color="#64748b"),
        )
        fig.update_layout(
            template=_PLOTLY_TEMPLATE,
            title="Convergence Trace :: Running Mean ±MCSE (top-6 most volatile)",
            height=720,
        )
        return fig

    def chart_rank_stability(self) -> go.Figure:
        """Kendall's W concordance + per-player rank swing (σ of rank).

        TUNING TAKEAWAY: W ≈ 1 means players are reordered identically
        across runs (deterministic ranking, possibly underwhelming).
        W close to 0 means each run scrambles the order (excessive RNG).
        The per-player rank swing shows WHERE reordering is concentrated:
        a player with swing 8 changes league position by ±8 slots run-to-run.

        STATS NOTE: this uses rank STANDARD DEVIATION (in rank units), not
        the previous rank coefficient of variation. CV = σ/μ is mechanically
        confounded with baseline, so CV falls
        with baseline by construction (r≈−0.9 here), consequently, not a
        finding. Rank σ is in interpretable units and free of that division
        artifact (though ranks are still bounded, so the very best/worst
        players can only swing inward). W's significance uses the large-sample
        χ² = m(n−1)W with df=n−1.
        """
        pivot = self.sim.pivot_table(
            index="PlayerID", columns="Run", values="Ovr"
        ).dropna()
        n_items, m_runs = pivot.shape
        if m_runs < 2 or n_items < 3:
            return _empty_fig("Rank Stability")

        ranks = pivot.rank(axis=0)
        W = kendalls_w(ranks.values)
        # Large-sample significance of W: χ² = m(n-1)W ~ χ²(n-1)
        chi2_stat = m_runs * (n_items - 1) * W
        p_w = float(scipy_stats.chi2.sf(chi2_stat, df=n_items - 1))
        swing = ranks.std(axis=1).rename("RankSwing")  # rank units

        fig = make_subplots(
            rows=1,
            cols=2,
            subplot_titles=(
                "<b>Rank-Swing Distribution (σ of rank)</b>",
                "<b>Rank Swing vs Baseline OVR</b>",
            ),
            horizontal_spacing=0.14,
        )
        fig.add_trace(
            go.Histogram(
                x=swing.values,
                nbinsx=30,
                marker_color="#2563eb",
                marker_line_width=0,
                opacity=0.85,
                showlegend=False,
                hovertemplate="rank swing %{x:.1f}<br>%{y} players<extra></extra>",
            ),
            row=1,
            col=1,
        )
        fig.add_vline(
            x=float(swing.median()),
            line_dash="dash",
            line_color="#dc2626",
            annotation_text=f"median {swing.median():.1f}",
            row=1,
            col=1,
        )

        ppl = self.ppl.set_index("PlayerID").join(swing, how="inner").reset_index()
        if not ppl.empty:
            fig.add_trace(
                go.Scatter(
                    x=ppl["Baseline"],
                    y=ppl["RankSwing"],
                    mode="markers",
                    marker=dict(color="#2563eb", size=8, opacity=0.7),
                    customdata=ppl[["Label", "Age"]].values,
                    hovertemplate=(
                        "<b>%{customdata[0]}</b><br>"
                        "Age %{customdata[1]}  |  base %{x:.0f}<br>"
                        "rank swing ±%{y:.1f} slots<extra></extra>"
                    ),
                    showlegend=False,
                ),
                row=1,
                col=2,
            )

        interp = (
            "rigid as RNG barely reorders players"
            if W > 0.85
            else "stable"
            if W > 0.7
            else "moderate"
            if W > 0.4
            else "unstable, the runs scramble the order"
        )
        sig = "significant" if p_w < 0.05 else "n.s."
        fig.add_annotation(
            text=(
                f"<b>Kendall's W = {W:.3f}</b> ({interp})  ·  "
                f"χ²({n_items - 1})={chi2_stat:.0f}, p={p_w:.2e} ({sig})  ·  "
                f"{n_items} players × {m_runs} runs"
            ),
            xref="paper",
            yref="paper",
            x=0.5,
            y=1.12,
            showarrow=False,
            font=dict(size=12, color="#0f172a"),
        )
        fig.update_xaxes(title_text="Per-player rank σ (slots)", row=1, col=1)
        fig.update_xaxes(title_text="Baseline OVR", row=1, col=2)
        fig.update_yaxes(title_text="Players", row=1, col=1)
        fig.update_yaxes(title_text="Rank swing (slots)", row=1, col=2)
        fig.update_layout(
            template=_PLOTLY_TEMPLATE,
            title="Rank Stability Across Monte Carlo Runs",
            height=500,
        )
        return fig

    # ════════════════════════════════════════════════════════════════════
    # §7  DIAGNOSTICS
    # ════════════════════════════════════════════════════════════════════
    def chart_outlier_detection(self) -> go.Figure:
        """Mahalanobis distance in (Mean Δ, Std Δ) space. Flags structural outliers.

        TUNING TAKEAWAY: Diamond points are players whose joint
        (mean, σ) signature is unlikely under the rest of the population.
        """
        ppl = self.ppl.dropna(subset=["MeanDelta", "StdDelta"])
        if len(ppl) < 4:
            return _empty_fig("Outlier Detection")
        X = ppl[["MeanDelta", "StdDelta"]].values
        dists = mahalanobis_distance(X)
        p = 1 - scipy_stats.chi2.cdf(dists**2, df=2)
        out = ppl.copy()
        out["Mahal"] = dists
        out["p"] = p
        sig = out[out["p"] < 0.05]
        normal = out[out["p"] >= 0.05]

        fig = go.Figure()
        fig.add_trace(
            go.Scatter(
                x=normal["MeanDelta"],
                y=normal["StdDelta"],
                mode="markers",
                marker=dict(color="#94a3b8", size=7, opacity=0.45),
                name="Within",
                hoverinfo="skip",
            )
        )
        if not sig.empty:
            fig.add_trace(
                go.Scatter(
                    x=sig["MeanDelta"],
                    y=sig["StdDelta"],
                    mode="markers+text",
                    marker=dict(
                        color="#dc2626",
                        size=12,
                        symbol="diamond",
                        line=dict(width=2, color="white"),
                    ),
                    text=sig["Label"].apply(lambda s: s.split("(")[0].strip()),
                    textposition="top center",
                    textfont=dict(size=9, color="#dc2626"),
                    customdata=sig[["Label", "Mahal", "p", "Age", "Baseline"]].values,
                    hovertemplate=(
                        "<b>%{customdata[0]}</b><br>"
                        "Age %{customdata[3]}  |  base %{customdata[4]:.0f}<br>"
                        "Mahal %{customdata[1]:.2f}  |  p=%{customdata[2]:.4f}<extra></extra>"
                    ),
                    name="Outlier (p<0.05)",
                )
            )
        # χ² 90%/99% ellipses, oriented by the actual covariance of (μΔ, σΔ).
        # An axis-aligned ellipse would disagree with the Mahalanobis flagging
        # whenever μΔ and σΔ are correlated, which they typically are.
        # https://carstenschelp.github.io/2018/09/14/Plot_Confidence_Ellipse_001.html
        cov = np.cov(X, rowvar=False)
        eigvals, eigvecs = np.linalg.eigh(cov)
        # Guard against tiny negative eigenvalues from float noise on near-singular cov
        eigvals = np.maximum(eigvals, 0.0)
        cx, cy = ppl["MeanDelta"].mean(), ppl["StdDelta"].mean()
        theta = np.linspace(0, 2 * np.pi, 120)
        unit_circle = np.column_stack([np.cos(theta), np.sin(theta)])
        transform = eigvecs @ np.diag(np.sqrt(eigvals))  # shape (2,2)

        for r_val, lbl, c in [
            (np.sqrt(scipy_stats.chi2.ppf(0.90, 2)), "90%", "#f97316"),
            (np.sqrt(scipy_stats.chi2.ppf(0.99, 2)), "99%", "#dc2626"),
        ]:
            ellipse = r_val * unit_circle @ transform.T  # (N, 2)
            fig.add_trace(
                go.Scatter(
                    x=cx + ellipse[:, 0],
                    y=cy + ellipse[:, 1],
                    mode="lines",
                    line=dict(color=c, width=2, dash="dash"),
                    name=f"χ² {lbl}",
                    hoverinfo="skip",
                )
            )
        fig.update_layout(
            template=_PLOTLY_TEMPLATE,
            title="Outlier Detection :: Mahalanobis Distance in (μΔ, σΔ) Space",
            xaxis_title="Mean OVR Δ",
            yaxis_title="Std OVR Δ",
        )
        return fig

    def chart_funnel(self) -> go.Figure:
        """σ(Δ) as a function of baseline OVR, a variance funnel.

        TUNING TAKEAWAY: If σ narrows toward high baselines you've got a
        working ceiling (less room to swing). If σ is flat or widens, the
        cap mechanism isn't biting and high-OVR players keep volatility
        their lower peers do too.
        """
        ppl = self.ppl.dropna(subset=["Baseline", "StdDelta"])
        if len(ppl) < 10:
            return _empty_fig("Funnel Plot")
        fig = go.Figure()
        fig.add_trace(
            go.Scatter(
                x=ppl["Baseline"],
                y=ppl["StdDelta"],
                mode="markers",
                marker=dict(
                    color=ppl["MeanDelta"],
                    colorscale="RdYlGn",
                    cmid=0,
                    size=8,
                    opacity=0.75,
                    line=dict(width=1, color="white"),
                    colorbar=dict(title="Mean Δ", thickness=15),
                    showscale=True,
                ),
                customdata=ppl[["Label", "Age", "PctPositive"]].values,
                hovertemplate=(
                    "<b>%{customdata[0]}</b><br>"
                    "Base %{x:.0f} | σΔ %{y:.1f}<br>"
                    "Age %{customdata[1]} | P(+) %{customdata[2]:.0%}<extra></extra>"
                ),
                name="Players",
            )
        )
        # Rolling window of σ to overlay
        try:
            ppl_s = ppl.sort_values("Baseline")
            win = max(5, len(ppl_s) // 10)
            roll_x = ppl_s["Baseline"].rolling(win, min_periods=3, center=True).mean()
            roll_y = ppl_s["StdDelta"].rolling(win, min_periods=3, center=True).mean()
            fig.add_trace(
                go.Scatter(
                    x=roll_x,
                    y=roll_y,
                    mode="lines",
                    name="Rolling mean σ",
                    line=dict(color="#dc2626", width=3),
                )
            )
            rho, p = scipy_stats.spearmanr(ppl["Baseline"], ppl["StdDelta"])
            interp = (
                "ceiling compresses variance"
                if rho < -0.3 and p < 0.05
                else "no compression detected"
                if p >= 0.05
                else "variance grows with baseline (unusual)"
            )
            fig.add_annotation(
                text=(
                    f"<b>Heteroscedasticity</b>: ρ={rho:+.2f} (p={p:.4f})<br>"
                    f"<b>Verdict</b>: {interp}"
                ),
                xref="paper",
                yref="paper",
                x=0.98,
                y=0.98,
                showarrow=False,
                font=dict(size=11, family="monospace", color="#0f172a"),
                bgcolor="rgba(255,255,255,0.92)",
                bordercolor="#e2e8f0",
                borderwidth=1,
                align="right",
            )
        except ValueError:
            pass
        fig.update_layout(
            template=_PLOTLY_TEMPLATE,
            title="Funnel Plot :: σ(Δ) vs Baseline OVR",
            xaxis_title="Baseline OVR",
            yaxis_title="Std of OVR Δ",
        )
        return fig

    # ════════════════════════════════════════════════════════════════════
    # §8  PLAYER EXPLORER  (interactive widget, not a go.Figure)
    # ════════════════════════════════════════════════════════════════════
    def build_player_explorer(self) -> str:
        """One embedded payload + a client-side selector that renders any
        single player's Δ distribution and OVR-weighted attribute movement.

        Efficiency: rather than emit 294 figures, we precompute a compact
        record per player.
        """
        sim = self.sim
        ppl = self.ppl
        if sim.empty or ppl.empty:
            return (
                '<div class="chart-card"><p style="color:#64748b">'
                "Player explorer unavailable (no data).</p></div>"
            )

        # Shared histogram edges over the robust global Δ range.
        d_all = sim["Delta"].dropna().values
        lo, hi = np.percentile(d_all, [0.5, 99.5])
        if hi <= lo:
            lo, hi = float(d_all.min()), float(d_all.max() + 1e-6)
        edges = np.linspace(lo, hi, 41)
        centers = ((edges[:-1] + edges[1:]) / 2).round(3)

        # Per-player Δ arrays in one groupby (avoids 294 filters).
        delta_by_player = {pid: g.values for pid, g in sim.groupby("PlayerID")["Delta"]}

        # Coef-weighted attribute movement from the per-player attr deltas.
        coef_map = dict(zip(Config.OVR_CALC_ORDER, Config.OVR_COEFFS))
        ad = self.ad
        avail_attrs = [
            a for a in self.cols.varying_attrs if not ad.empty and a in ad.columns
        ]

        league_mean = float(ppl["MeanDelta"].mean())

        players = []
        # Sort by |MeanDelta| desc so the default selection is interesting.
        ppl_sorted = ppl.reindex(
            ppl["MeanDelta"].abs().sort_values(ascending=False).index
        )
        for _, row in ppl_sorted.iterrows():
            pid = row["PlayerID"]
            dvals = delta_by_player.get(pid)
            if dvals is None or len(dvals) == 0:
                continue
            counts, _ = np.histogram(dvals, bins=edges)
            rec = {
                "id": int(pid) if str(pid).isdigit() else str(pid),
                "label": str(row.get("Label", pid)),
                "age": int(row["Age"]) if pd.notnull(row["Age"]) else None,
                "base": round(float(row["Baseline"]), 1)
                if pd.notnull(row["Baseline"])
                else None,
                "mean": round(float(row["MeanDelta"]), 2),
                "std": round(float(row["StdDelta"]), 2)
                if pd.notnull(row["StdDelta"])
                else 0.0,
                "p05": round(float(row["P05_Delta"]), 2),
                "p50": round(float(row["P50_Delta"]), 2),
                "p95": round(float(row["P95_Delta"]), 2),
                "pct": round(float(row["PctPositive"]), 3),
                "hist": counts.astype(int).tolist(),
            }
            if avail_attrs and pid in ad.index:
                contrib = {
                    a: round(float(ad.at[pid, a]) * coef_map.get(a, 0.0), 4)
                    for a in avail_attrs
                }
                rec["attrs"] = contrib
            else:
                rec["attrs"] = {}
            players.append(rec)

        group_colors = {
            a: GROUP_COLORS.get(_attr_group(a), "#475569") for a in avail_attrs
        }

        payload = {
            "centers": centers.tolist(),
            "leagueMean": round(league_mean, 3),
            "groupColors": group_colors,
            "players": players,
        }
        data_json = json.dumps(payload, separators=(",", ":")).replace("</", "<\\/")

        # Widget markup + script. The <script> executes on page load (it is
        # written statically into the file, not injected via innerHTML).
        widget = f"""<div class="chart-card">
  <div class="explorer-controls" style="display:flex;gap:12px;align-items:center;
       flex-wrap:wrap;margin-bottom:14px;">
    <label for="explorer-select" style="font-weight:600;color:#334155;">Player</label>
    <input list="explorer-options" id="explorer-select" placeholder="type to search…"
           style="flex:1;min-width:240px;max-width:420px;padding:8px 10px;
           border:1px solid #e2e8f0;border-radius:8px;font-size:0.95rem;">
    <datalist id="explorer-options"></datalist>
    <span id="explorer-stats" style="color:#64748b;font-size:0.85rem;"></span>
  </div>
  <div style="display:grid;grid-template-columns:1fr 1fr;gap:16px;">
    <div id="explorer-dist" style="min-height:360px;"></div>
    <div id="explorer-attrs" style="min-height:360px;"></div>
  </div>
  <script type="application/json" id="explorer-data">{data_json}</script>
  <script type="text/javascript">{_EXPLORER_JS}</script>
</div>"""
        return widget


# Client-side player-explorer logic (plain string; runs on page load).
_EXPLORER_JS = """
(function () {
  var raw = document.getElementById("explorer-data");
  if (!raw) return;
  var D = JSON.parse(raw.textContent);
  var byLabel = {};
  D.players.forEach(function (p) { byLabel[p.label] = p; });

  var input = document.getElementById("explorer-select");
  var list = document.getElementById("explorer-options");
  var stats = document.getElementById("explorer-stats");
  D.players.forEach(function (p) {
    var o = document.createElement("option");
    o.value = p.label;
    list.appendChild(o);
  });

  var TEMPLATE_FONT = {family: "Inter, Roboto, sans-serif", color: "#1e293b"};

  function draw(p) {
    if (!p) return;
    stats.textContent = "age " + p.age + " · base " + p.base +
      " · μΔ " + (p.mean >= 0 ? "+" : "") + p.mean +
      " · σ " + p.std + " · P(+) " + Math.round(p.pct * 100) + "%";

    // Distribution
    var distTrace = {
      type: "bar", x: D.centers, y: p.hist,
      marker: {color: "#2563eb", opacity: 0.85},
      hovertemplate: "Δ %{x:.2f}<br>%{y} runs<extra></extra>"
    };
    var shapes = [
      {type: "line", x0: p.mean, x1: p.mean, yref: "paper", y0: 0, y1: 1,
       line: {color: "#dc2626", width: 2, dash: "dash"}},
      {type: "line", x0: p.p05, x1: p.p05, yref: "paper", y0: 0, y1: 1,
       line: {color: "#f97316", width: 1.5, dash: "dot"}},
      {type: "line", x0: p.p95, x1: p.p95, yref: "paper", y0: 0, y1: 1,
       line: {color: "#f97316", width: 1.5, dash: "dot"}},
      {type: "line", x0: D.leagueMean, x1: D.leagueMean, yref: "paper", y0: 0, y1: 1,
       line: {color: "#94a3b8", width: 1.5, dash: "dashdot"}}
    ];
    var distLayout = {
      title: {text: "OVR Δ distribution across runs", font: {size: 15}},
      font: TEMPLATE_FONT, paper_bgcolor: "#fff", plot_bgcolor: "#f8fafc",
      margin: {l: 55, r: 20, t: 50, b: 45}, shapes: shapes,
      xaxis: {title: "OVR Δ", gridcolor: "#e2e8f0", zerolinecolor: "#94a3b8"},
      yaxis: {title: "runs", gridcolor: "#e2e8f0"},
      annotations: [
        {x: p.mean, yref: "paper", y: 1.02, text: "μ", showarrow: false,
         font: {color: "#dc2626", size: 12}},
        {x: D.leagueMean, yref: "paper", y: 1.02, text: "league μ",
         showarrow: false, font: {color: "#94a3b8", size: 10}}
      ]
    };
    Plotly.newPlot("explorer-dist", [distTrace], distLayout,
                   {responsive: true, displayModeBar: false});

    // Attribute contributions (coef-weighted), sorted ascending for a bar
    var keys = Object.keys(p.attrs || {});
    keys.sort(function (a, b) { return p.attrs[a] - p.attrs[b]; });
    var xs = keys.map(function (k) { return p.attrs[k]; });
    var colors = keys.map(function (k) { return D.groupColors[k] || "#475569"; });
    var attrTrace = {
      type: "bar", orientation: "h", y: keys, x: xs,
      marker: {color: colors, opacity: 0.9},
      hovertemplate: "%{y}<br>OVR-weighted Δ %{x:+.3f}<extra></extra>"
    };
    var attrLayout = {
      title: {text: "Attribute movement (coef × mean Δ)", font: {size: 15}},
      font: TEMPLATE_FONT, paper_bgcolor: "#fff", plot_bgcolor: "#f8fafc",
      margin: {l: 55, r: 20, t: 50, b: 45},
      xaxis: {title: "OVR points contributed", gridcolor: "#e2e8f0",
              zerolinecolor: "#1e293b", zerolinewidth: 1.5},
      yaxis: {gridcolor: "#e2e8f0"}
    };
    Plotly.newPlot("explorer-attrs", [attrTrace], attrLayout,
                   {responsive: true, displayModeBar: false});
  }

  input.addEventListener("change", function () {
    var p = byLabel[input.value];
    if (p) draw(p);
  });
  // Default to the first (most extreme μΔ) player.
  if (D.players.length) {
    input.value = D.players[0].label;
    draw(D.players[0]);
  }
})();
"""


def _validate_chart_registry() -> None:
    missing = [
        f"{sec['id']}/{name}"
        for sec in SECTIONS
        for name in sec["charts"]
        if name not in WIDGET_CHARTS and not callable(getattr(ChartBuilder, name, None))
    ]
    if missing:
        raise RuntimeError(f"SECTIONS references unknown chart methods: {missing}")


# LAZY-RENDER HTML : self-contained, embed-friendly, instant open.
#
# Every figure is serialized to JSON and dropped into a <script
# type="application/json"> tag. An IntersectionObserver renders each chart
# with Plotly.newPlot only when it scrolls near the viewport. A 30-chart
# dashboard therefore opens instantly instead of paying the full render
# cost up front, and the file stays a single self-contained .html (plotly.js
# inlined, no network fetches) that embeds cleanly via <iframe> / srcdoc.

_LAZY_JS = """
(function () {
  function renderChart(host) {
    if (host.dataset.rendered) return;
    host.dataset.rendered = "1";
    var payload = document.getElementById(host.dataset.src);
    if (!payload) return;
    var fig;
    try { fig = JSON.parse(payload.textContent); }
    catch (e) { host.innerHTML = "<p style='color:#dc2626'>chart parse error</p>"; return; }
    var cfg = Object.assign({responsive: true, displModeBar: false}, fig.config || {});
    Plotly.newPlot(host, fig.data, fig.layout, cfg);
  }
  var obs = new IntersectionObserver(function (entries) {
    entries.forEach(function (e) {
      if (e.isIntersecting) { renderChart(e.target); obs.unobserve(e.target); }
    });
  }, {rootMargin: "800px 0px"});
  document.querySelectorAll(".chart-host").forEach(function (h) { obs.observe(h); });
  // Re-layout visible charts on window resize (responsive:true handles size,
  // but a debounced nudge keeps subplot heights honest).
  var t = null;
  window.addEventListener("resize", function () {
    clearTimeout(t);
    t = setTimeout(function () {
      document.querySelectorAll(".chart-host[data-rendered]").forEach(function (h) {
        Plotly.Plots.resize(h);
      });
    }, 200);
  });
})();
"""


def _fig_layout_height(fig: go.Figure, default: int = 450) -> int:
    """Best-effort pixel height so the placeholder reserves space (no scroll jump)."""
    try:
        h = fig.layout.height
        if h:
            return int(h)
    except Exception:
        pass
    return default


def _fig_to_payload(fig: go.Figure) -> str:
    """Figure -> JSON string safe to inline inside a <script> tag."""
    # to_json emits compact, plotly-native JSON. Escaping </ prevents an
    # embedded "</script>" in any label/hovertext from closing the tag early.
    return fig.to_json().replace("</", "<\\/")


def _render_html(
    path: Path,
    title: str,
    subtitle: str,
    hero_h1: str,
    hero_p: str,
    nav: List[Tuple[str, str]],
    stat_cards: List[Tuple[str, str, Optional[str]]],
    sections: List[Dict],
    charts: Dict[str, go.Figure],
) -> None:
    """Shared self-contained lazy-render HTML writer.

    nav          : [(anchor_id, label), ...]
    stat_cards   : [(label, value, color_or_None), ...]
    sections     : [{'id','title','intro','charts':[chart_key,...]}, ...]
                   chart_key indexes into `charts`.
    charts       : {chart_key: go.Figure}
    """
    logger.info(f"Rendering HTML dashboard --> {path}")
    plotly_js = get_plotlyjs()

    nav_html = "".join(f'<a href="#{aid}">{lbl}</a>' for aid, lbl in nav)

    cards_html = []
    for lbl, val, color in stat_cards:
        style = f' style="color:{color}"' if color else ""
        cards_html.append(
            f'<div class="stat-card"><div class="lbl">{lbl}</div>'
            f'<div class="val"{style}>{val}</div></div>'
        )

    payloads: List[str] = []
    body_parts: List[str] = []
    counter = 0
    for sec in sections:
        inner = []
        for ckey in sec["charts"]:
            # build_all() keys charts as "<section_id>/<method>"; SECTIONS
            # list bare method names. Resolve the qualified key first, then
            # fall back to a bare-name lookup for robustness.
            fig = charts.get(f"{sec['id']}/{ckey}") or charts.get(ckey)
            if fig is None:
                continue
            # Widget charts arrive as ready-to-inject HTML strings (they carry
            # their own <script>); emit them directly, no lazy JSON payload.
            if isinstance(fig, str):
                inner.append(fig)
                continue
            counter += 1
            pid = f"payload-{counter}"
            hid = f"host-{counter}"
            height = _fig_layout_height(fig)
            payloads.append(
                f'<script type="application/json" id="{pid}">'
                f"{_fig_to_payload(fig)}</script>"
            )
            inner.append(
                f'<div class="chart-card">'
                f'<div class="chart-host" id="{hid}" data-src="{pid}" '
                f'style="min-height:{height}px"></div>'
                f"</div>"
            )
        body_parts.append(
            f'<section id="{sec["id"]}" class="section">'
            f'  <h2 class="section-title">{sec["title"]}</h2>'
            f'  <p class="section-intro">{sec["intro"]}</p>'
            f"  {''.join(inner)}"
            f"</section>"
        )

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>{title}</title>
    <script type="text/javascript">{plotly_js}</script>
    <style>
        :root {{
            --bg: #f1f5f9; --card: #ffffff; --text: #0f172a;
            --muted: #64748b; --border: #e2e8f0; --accent: #2563eb;
        }}
        * {{ box-sizing: border-box; }}
        body {{
            font-family: 'Inter', -apple-system, BlinkMacSystemFont, "Segoe UI",
                         Roboto, sans-serif;
            background: var(--bg); color: var(--text);
            margin: 0; padding: 0;
        }}
        .topbar {{
            position: sticky; top: 0; z-index: 100;
            background: rgba(255,255,255,0.96); backdrop-filter: blur(8px);
            border-bottom: 1px solid var(--border);
            padding: 12px 24px; display: flex; flex-wrap: wrap;
            align-items: center; gap: 12px;
        }}
        .topbar .logo {{
            font-weight: 800; font-size: 1.1rem; letter-spacing: -0.02em;
            margin-right: 16px;
        }}
        .topbar a {{
            color: var(--muted); text-decoration: none; font-size: 0.85rem;
            padding: 6px 10px; border-radius: 6px; transition: all 0.15s;
        }}
        .topbar a:hover {{ background: var(--bg); color: var(--text); }}
        .hero {{
            max-width: 1400px; margin: 0 auto; padding: 50px 24px 20px;
        }}
        .hero h1 {{
            font-size: 2.4rem; font-weight: 800; letter-spacing: -0.025em;
            margin: 0 0 6px 0;
        }}
        .hero p {{ color: var(--muted); margin: 0; font-size: 1.05rem; }}
        .stats-grid {{
            max-width: 1400px; margin: 24px auto 40px; padding: 0 24px;
            display: grid; grid-template-columns: repeat(auto-fit, minmax(180px, 1fr));
            gap: 14px;
        }}
        .stat-card {{
            background: var(--card); border: 1px solid var(--border);
            border-radius: 10px; padding: 16px 20px;
        }}
        .stat-card .lbl {{
            font-size: 0.75rem; text-transform: uppercase; letter-spacing: 0.06em;
            color: var(--muted); margin-bottom: 4px;
        }}
        .stat-card .val {{
            font-size: 1.6rem; font-weight: 700; color: var(--text);
        }}
        .section {{
            max-width: 1400px; margin: 0 auto; padding: 40px 24px;
            border-top: 1px solid var(--border);
        }}
        .section-title {{
            font-size: 1.6rem; font-weight: 700; margin: 0 0 10px 0;
            letter-spacing: -0.02em;
        }}
        .section-intro {{
            color: var(--muted); font-size: 0.95rem; line-height: 1.6;
            margin: 0 0 28px 0; max-width: 900px;
        }}
        .chart-card {{
            background: var(--card); border: 1px solid var(--border);
            border-radius: 12px; padding: 24px; margin-bottom: 22px;
            box-shadow: 0 1px 3px rgba(0,0,0,0.03);
            transition: box-shadow 0.2s;
        }}
        .chart-card:hover {{ box-shadow: 0 4px 12px rgba(0,0,0,0.06); }}
        .chart-host {{ width: 100%; }}
        footer {{
            text-align: center; color: var(--muted); font-size: 0.8rem;
            padding: 30px 0 50px;
        }}
        footer a {{
            color: var(--muted); text-decoration: none; font-weight: 500;
            transition: color 0.15s;
        }}
        footer a:hover {{ color: var(--accent); }}
    </style>
</head>
<body>
    <nav class="topbar">
        <span class="logo">Progbox</span>
        {nav_html}
    </nav>
    <div class="hero">
        <h1>{hero_h1}</h1>
        <p>{hero_p}</p>
    </div>
    <div class="stats-grid">
        {"".join(cards_html)}
    </div>
    {"".join(body_parts)}
    {"".join(payloads)}
    <footer>
        <div>
            Generated by <strong>Progbox</strong> &middot;
            <time datetime="{time.strftime("%Y-%m-%d")}">{time.strftime("%Y-%m-%d %H:%M")}</time>
            &middot; {subtitle}
        </div>
        <div>
            &copy; {time.strftime("%Y")}
            <a href="https://github.com/akshayexists" target="_blank" rel="noopener noreferrer">@akshayexists</a>
        </div>
    </footer>
    <script type="text/javascript">{_LAZY_JS}</script>
</body>
</html>
"""
    path.write_text(html, encoding="utf-8")
    logger.info(f"HTML saved · {path.stat().st_size / 1e6:.2f} MB")


class HTMLDashboard:
    """Single-run dashboard (thin wrapper over the shared lazy renderer)."""

    def __init__(self, charts: Dict[str, go.Figure], ds: Dataset, label: str = ""):
        self._charts = charts
        self._ds = ds
        self._label = label

    def render(self, path: Path) -> None:
        conv = self._ds.convergence
        pct = conv["pct_converged"]
        max_mcse = conv["max_mcse"]
        n_players = self._ds.sim["PlayerID"].nunique()
        n_runs = self._ds.sim["Run"].nunique()
        n_inputs = len(self._ds.cols.inputs)
        conv_color = (
            "#16a34a" if pct >= 0.95 else "#f97316" if pct >= 0.8 else "#dc2626"
        )

        subtitle = f"script: {self._label}" if self._label else "single-run analysis"
        _render_html(
            path,
            title="Progbox · Monte Carlo Tuning Dashboard",
            subtitle=subtitle,
            hero_h1="Monte Carlo Tuning Dashboard",
            hero_p=(
                "Diagnostic suite for progression scripts. Each section "
                "answers a different class of tuning question. Read "
                "top-down or jump via the nav above."
            ),
            nav=[(sec["id"], sec["title"]) for sec in SECTIONS],
            stat_cards=[
                ("Players", f"{n_players:,}", None),
                ("Runs", f"{n_runs:,}", None),
                ("Model Inputs", f"{n_inputs}", None),
                ("Convergence", f"{pct:.0%}", conv_color),
                ("Max MCSE", f"{max_mcse:.2f}", None),
            ],
            sections=SECTIONS,
            charts=self._charts,
        )


# EXCEL WORKBOOK
class ExcelWorkbook:
    """Multi-sheet Excel artifact mirroring the dashboard's data.

    PERF: sheets are written with ws.append(row) over pre-converted rows
    (NaN-->None done once, vectorized) instead of cell-by-cell ws.cell().
    """

    def __init__(self, ds: Dataset, full_excel: bool = False):
        self._ds = ds
        self._full = full_excel

    def render(self, path: Path) -> None:
        logger.info(f"Writing Excel workbook --> {path}")
        wb = Workbook()
        self._sheet_players(wb)
        self._sheet_per_age(wb)
        self._sheet_attributes(wb)
        self._sheet_attr_progression(wb)
        self._sheet_inputs(wb)
        self._sheet_icc(wb)
        self._sheet_convergence(wb)
        self._sheet_outliers(wb)
        self._sheet_team_summary(wb)
        if self._full:
            self._sheet_all_runs(wb)
        if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
            del wb["Sheet"]
        wb.active = wb.sheetnames.index("Players")
        wb.save(path)
        logger.info(f"Excel saved · {path.stat().st_size / 1e6:.2f} MB")

    # helpers
    def _write(self, wb: Workbook, title: str, df: pd.DataFrame) -> None:
        if "Sheet" in wb.sheetnames and len(wb.sheetnames) == 1:
            ws = wb["Sheet"]
            ws.title = title
        else:
            ws = wb.create_sheet(title)

        # PERF: append whole rows. Convert NaN-->None once (openpyxl writes
        # None as an empty cell; a raw float nan would serialize as "NaN").
        if not df.empty:
            clean = df.astype(object).where(pd.notnull(df), None)
            ws.append(list(df.columns))
            for row in clean.itertuples(index=False, name=None):
                ws.append(list(row))

        n_cols = len(df.columns)
        n_rows = len(df) + 1
        if n_cols > 0 and n_rows > 1:
            ws.auto_filter.ref = f"A1:{get_column_letter(n_cols)}{n_rows}"
        # Auto-size first 12 columns roughly
        for c in range(1, min(n_cols + 1, 13)):
            ws.column_dimensions[get_column_letter(c)].width = 14

    # sheets
    def _sheet_players(self, wb: Workbook) -> None:
        self._write(wb, "Players", self._ds.per_player.copy().round(4))
        logger.info("  · Players")

    def _sheet_per_age(self, wb: Workbook) -> None:
        self._write(wb, "Per Age", self._ds.per_age.copy().round(4))
        logger.info("  · Per Age")

    def _sheet_attributes(self, wb: Workbook) -> None:
        ad = self._ds.attr_deltas
        if ad.empty:
            self._write(wb, "Attributes", pd.DataFrame())
            logger.info("  · Attributes [empty]")
            return
        ppl = self._ds.per_player
        avail = [c for c in self._ds.cols.varying_attrs if c in ad.columns]
        d = ad[avail].round(2).rename(columns=lambda c: f"D_{c}")
        m = self._ds.sim.groupby("PlayerID")[avail].mean().round(2)
        joined = (
            ppl[["PlayerID", "Name", "Team", "Age", "Baseline"]]
            .set_index("PlayerID")
            .join(d)
            .join(m)
            .reset_index()
            .sort_values(["Age", "Baseline"], ascending=[True, False])
        )
        self._write(wb, "Attributes", joined)
        logger.info("  · Attributes")

    def _sheet_attr_progression(self, wb: Workbook) -> None:
        ad = self._ds.attr_deltas
        if ad.empty:
            self._write(wb, "Attr Progression", pd.DataFrame())
            logger.info("  · Attr Progression [empty]")
            return
        ad2 = ad.copy()
        ad2["AgeTier"] = Dataset.assign_tiers(ad2["Age"])
        avail = [c for c in self._ds.cols.varying_attrs if c in ad2.columns]
        rows = []
        for tier in AGE_COLORS:
            sub = ad2[ad2["AgeTier"] == tier]
            if sub.empty:
                continue
            for attr in avail:
                vals = sub[attr].dropna()
                if not len(vals):
                    continue
                _, lo, hi = bootstrap_ci(
                    vals.values, n_bootstrap=1500, seed=hash((tier, attr)) & 0x7FFFFFFF
                )
                rows.append(
                    {
                        "AgeTier": tier,
                        "Attribute": attr,
                        "Group": _attr_group(attr),
                        "N": len(vals),
                        "Mean": vals.mean(),
                        "SE": vals.sem(),
                        "CI_lo": lo,
                        "CI_hi": hi,
                    }
                )
        self._write(wb, "Attr Progression", pd.DataFrame(rows).round(4))
        logger.info("  · Attr Progression")

    def _sheet_inputs(self, wb: Workbook) -> None:
        """Model-input correlation matrix vs major outcome metrics."""
        ppl = self._ds.per_player
        inputs = [
            c for c in self._ds.cols.inputs if c in ppl.columns and ppl[c].std() > 0
        ]
        if not inputs:
            self._write(wb, "Inputs", pd.DataFrame())
            logger.info("  · Inputs [empty]")
            return
        metrics = [
            "MeanDelta",
            "StdDelta",
            "P05_Delta",
            "P95_Delta",
            "PctPositive",
            "IQR",
            "Sharpe",
        ]
        rows = []
        for inp in inputs:
            row: Dict = {"Input": inp, "Mean": ppl[inp].mean(), "Std": ppl[inp].std()}
            for m in metrics:
                r, p = scipy_stats.spearmanr(ppl[inp], ppl[m])
                row[f"{m}_rho"] = r
                row[f"{m}_p"] = p
            rows.append(row)
        self._write(wb, "Inputs", pd.DataFrame(rows).round(4))
        logger.info("  · Inputs")

    def _sheet_icc(self, wb: Workbook) -> None:
        self._write(wb, "ICC", self._ds.icc.round(4))
        logger.info("  · ICC")

    def _sheet_convergence(self, wb: Workbook) -> None:
        c = self._ds.convergence["stats"].reset_index().round(4)
        c["Converged"] = c["MCSE"] < Config.MCSE_THRESHOLD
        self._write(wb, "Convergence", c)
        logger.info("  · Convergence")

    def _sheet_outliers(self, wb: Workbook) -> None:
        ppl = self._ds.per_player.dropna(subset=["MeanDelta", "StdDelta"])
        if len(ppl) < 4:
            self._write(wb, "Outliers", pd.DataFrame())
            logger.info("  · Outliers [empty]")
            return
        X = ppl[["MeanDelta", "StdDelta"]].values
        dists = mahalanobis_distance(X)
        out = ppl.copy()
        out["Mahal"] = dists
        out["p"] = 1 - scipy_stats.chi2.cdf(dists**2, df=2)
        out = out.sort_values("Mahal", ascending=False)
        cols = [
            "PlayerID",
            "Label",
            "Age",
            "Baseline",
            "MeanDelta",
            "StdDelta",
            "P05_Delta",
            "P95_Delta",
            "Mahal",
            "p",
        ]
        self._write(wb, "Outliers", out[cols].round(4))
        logger.info("  · Outliers")

    def _sheet_team_summary(self, wb: Workbook) -> None:
        team = (
            self._ds.per_player.groupby("Team")["MeanDelta"]
            .agg(
                Players="count",
                Mean="mean",
                Std="std",
                Min="min",
                Max="max",
                PctImproved=lambda s: (s > 0).mean(),
            )
            .round(3)
            .sort_values("Mean", ascending=False)
            .reset_index()
        )
        self._write(wb, "Team Summary", team)
        logger.info("  · Team Summary")

    def _sheet_all_runs(self, wb: Workbook) -> None:
        raw = (
            self._ds.sim.sort_values(["PlayerID", "Run"])
            .reset_index(drop=True)
            .round(4)
        )
        self._write(wb, "All Runs", raw)
        logger.info("  · All Runs [full]")


# COMPARISON MODE : head-to-head across N progression scripts.
#
# The philosophy shift: the single-run dashboard answers "is THIS script
# healthy?", this focuses more on seeing which script produces a better
#   · peak age & decline slope  (does the age curve land where I want?)
#   · prime-age talent separation (do studs out-progress scrubs at 25-27?)
#   · league drift              (systemic inflation/deflation)
#   · run-to-run noise / ICC    (signal vs RNG)
#   · ceiling behaviour         (P99 OVR, % player-runs over cap: Wall check)
#   · god-prog rate             (rare-flat-boost frequency)
# Every one of these is distilled to a single number per script in the
# SCORECARD, then the overlay charts show the shape behind each number.


class RunSummary:
    """Per-run-directory KPI bundle. Loads sim, computes summaries, then
    releases the heavy frames. Only compact per-age / per-player / attr
    tables and scalar KPIs are retained so N runs stay memory-cheap."""

    def __init__(self, run_dir: Path, ceiling: float):
        self.run_dir = run_dir
        self.ceiling = ceiling
        loader = DataLoader()
        raw = run_dir / "raw"
        sim = loader.load_sim(raw)
        baseline = loader.load_baseline(run_dir / "raw")
        meta = loader.load_metadata(run_dir)

        prog = meta.get("progression", {}) if isinstance(meta, dict) else {}
        self.label = prog.get("name") or prog.get("id") or run_dir.name
        self.n_godprogs = loader.load_godprogs(run_dir)

        ds = Dataset(sim, baseline)
        self.n_players = int(sim["PlayerID"].nunique())
        self.n_runs = int(sim["Run"].nunique())

        # Retained compact tables (used by overlay charts)
        self.per_age = ds.per_age.copy()
        self.per_player = ds.per_player.copy()
        self.attr_deltas = ds.attr_deltas.copy()
        self.icc = ds.icc.copy()
        self.varying_attrs = list(ds.cols.varying_attrs)

        # ── scalar KPIs ────────────────────────────────────────────────
        kpi = estimate_peak(self.per_age)
        self.peak_age = kpi["peak_age"]
        self.decline_slope = kpi["decline_slope"]

        self.drift = float(self.per_player["MeanDelta"].mean())
        self.median_sigma = float(self.per_player["StdDelta"].median())

        icc_ovr = ds.icc[ds.icc["Attribute"] == "Ovr (agg)"]["ICC"]
        self.icc_ovr = float(icc_ovr.iloc[0]) if len(icc_ovr) else np.nan
        attr_icc = ds.icc[ds.icc["Attribute"] != "Ovr (agg)"]["ICC"]
        self.icc_attr_median = float(attr_icc.median()) if len(attr_icc) else np.nan

        # Kendall's W on a run subsample (full pivot is expensive & redundant)
        self.kendall_w = self._kendall_w(sim)

        # Ceiling behaviour
        self.p99_ovr = float(np.nanpercentile(sim["Ovr"].values, 99))
        self.max_ovr = float(sim["Ovr"].max())
        self.pct_over_ceiling = float((sim["Ovr"].values > ceiling).mean())

        self.godprogs_per_run = self.n_godprogs / self.n_runs if self.n_runs else 0.0

        # Prime + banded separation gaps (multiplayer-fairness KPI)
        self.prod_label = production_series(self.per_player)[1]
        lo, hi = Config.PRIME_AGES
        self.prime_sep = separation_gap(self.per_player, lo, hi)[0]
        # OVR-controlled version: isolates 'talent separates at equal OVR'
        # from the ceiling merely capping high-OVR studs.
        self.prime_sep_adj = partial_separation_gap(self.per_player, lo, hi)[0]
        self.sep_bands = {
            f"{a}-{b}": separation_gap(self.per_player, a, b)[0]
            for a, b in Config.SEP_BANDS
        }

        del ds, sim, baseline

    @staticmethod
    def _kendall_w(sim: pd.DataFrame) -> float:
        runs = sim["Run"].unique()
        if len(runs) > Config.KENDALL_MAX_RUNS:
            rng = np.random.default_rng(0)
            runs = rng.choice(runs, Config.KENDALL_MAX_RUNS, replace=False)
        pivot = (
            sim[sim["Run"].isin(runs)]
            .pivot_table(index="PlayerID", columns="Run", values="Ovr")
            .dropna()
        )
        if pivot.shape[1] < 2 or pivot.shape[0] < 3:
            return np.nan
        return float(kendalls_w(pivot.rank(axis=0).values))

    def scorecard_row(self) -> Dict:
        return {
            "Script": self.label,
            "Players": self.n_players,
            "Runs": self.n_runs,
            "PeakAge": self.peak_age,
            "PrimeSep": self.prime_sep,
            "PrimeSep(OVRadj)": self.prime_sep_adj,
            "DeclineSlope": self.decline_slope,
            "Drift": self.drift,
            "MedianσΔ": self.median_sigma,
            "ICC(Ovr)": self.icc_ovr,
            "KendallW": self.kendall_w,
            "P99 Ovr": self.p99_ovr,
            "%>cap": self.pct_over_ceiling,
            "GodProg/run": self.godprogs_per_run,
        }


CMP_SECTIONS: List[Dict] = [
    {
        "id": "scorecard",
        "title": "§1 · Scorecard",
        "intro": "Every tuning KPI distilled to one number per script. Peak "
        "age and decline slope come from the shared zero-crossing "
        "estimator (identical to each single-run age curve). "
        "PrimeSep is the raw mean-Δ gap between top and bottom "
        "production terciles at ages 25-27. PrimeSep(OVRadj) calculates "
        "this same gap within specific Overall (OVR) tiers"
        "it controls for OVR level, so a soft ceiling capping "
        "high-OVR studs won't masquerade as unfairness. Read them "
        "together: raw ≪ OVRadj means the ratings ceiling is "
        "compressing the top; both negative means model is unfair. ",
        "charts": ["cmp_scorecard"],
    },
    {
        "id": "age",
        "title": "§2 · Age Curve",
        "intro": "The single most consequential property, overlaid. Look at "
        "where each curve crosses zero (peak) and how steeply it "
        "falls after. Bands are 95% CI; separated bands mean the "
        "scripts differ beyond run noise.",
        "charts": ["cmp_age_curve", "cmp_sigma_by_age"],
    },
    {
        "id": "separation",
        "title": "§3 · Talent Separation",
        "intro": "Do better players progress more, at the same age? Each "
        "band shows the top-minus-bottom production-tercile Δ gap. "
        "For multiplayer fairness you want this clearly positive at "
        "prime age and fading with age (talent develops the young, "
        "not the old).",
        "charts": ["cmp_separation_bands", "cmp_delta_vs_baseline"],
    },
    {
        "id": "ceiling",
        "title": "§4 · Ceiling & Outcomes",
        "intro": "How each script behaves against the soft cap. The "
        "improvement-probability curves show where the cap starts "
        "biting; the Δ-vs-baseline overlay (previous section) shows "
        "the same in level terms.",
        "charts": ["cmp_improve_probability"],
    },
    {
        "id": "noise",
        "title": "§5 · Noise & Determinism",
        "intro": "Run-to-run spread and how much of the outcome is player "
        "identity vs RNG. Tighter StdDelta KDE = less noisy; higher "
        "ICC = more deterministic. You want enough noise to be fun, "
        "not so much that identity stops mattering.",
        "charts": ["cmp_noise_kde", "cmp_icc_bars"],
    },
    {
        "id": "fingerprint",
        "title": "§6 · Attribute Fingerprint",
        "intro": "Which attributes each script actually moves, weighted by "
        "their OVR coefficient so the bars are in OVR-impact units. "
        "Two scripts with the same age curve can get there through "
        "very different attribute movement",
        "charts": ["cmp_attr_fingerprint"],
    },
]


class ComparisonBuilder:
    """Overlay charts across a list of RunSummary objects."""

    def __init__(self, runs: List[RunSummary]):
        self._runs = runs
        self._colors = {
            r.label: SCRIPT_PALETTE[i % len(SCRIPT_PALETTE)] for i, r in enumerate(runs)
        }

    def build_all(self) -> Dict[str, go.Figure]:
        out: Dict[str, go.Figure] = {}
        for sec in CMP_SECTIONS:
            for name in sec["charts"]:
                key = f"{sec['id']}/{name}"
                try:
                    out[key] = getattr(self, name)()
                except Exception as e:
                    logger.error(
                        f"  cmp chart {name} FAILED: {type(e).__name__}: {e}",
                        exc_info=True,
                    )
                    out[key] = _empty_fig(name, f"ERROR: {type(e).__name__}")
        return out

    def scorecard_df(self) -> pd.DataFrame:
        return pd.DataFrame([r.scorecard_row() for r in self._runs])

    # ── §1 scorecard ──────────────────────────────────────────────────
    def cmp_scorecard(self) -> go.Figure:
        """Head-to-head KPI table. One column per script.

        TUNING TAKEAWAY: the whole tuning loop on one screen. Read PeakAge
        and DeclineSlope against intent, PrimeSep for fairness, %>cap for
        the ceiling, ICC/KendallW for the signal-vs-noise balance.
        """
        df = self.scorecard_df()
        # Transpose: metrics as rows, scripts as columns (easier to compare)
        metrics = [c for c in df.columns if c != "Script"]
        header = ["Metric"] + df["Script"].tolist()

        def fmt(metric, val):
            if pd.isna(val):
                return "-"
            if metric in ("Players", "Runs"):
                return f"{int(val):,}"
            if metric in ("%>cap",):
                return f"{val:.2%}"
            if metric in ("Drift", "PrimeSep", "PrimeSep(OVRadj)", "DeclineSlope"):
                return f"{val:+.3f}"
            return f"{val:.2f}"

        cells = [metrics]
        colors_per_script = []
        for _, row in df.iterrows():
            cells.append([fmt(m, row[m]) for m in metrics])
            colors_per_script.append(self._colors[row["Script"]])

        # Header fill: neutral for Metric col, script color tint for the rest
        header_fill = ["#0f172a"] + colors_per_script
        header_font = ["white"] + ["white"] * len(colors_per_script)

        fig = go.Figure(
            data=[
                go.Table(
                    columnwidth=[130] + [110] * len(df),
                    header=dict(
                        values=[f"<b>{h}</b>" for h in header],
                        fill_color=header_fill,
                        font=dict(color=header_font, size=12),
                        align="left",
                        height=32,
                    ),
                    cells=dict(
                        values=cells,
                        fill_color=[["#f8fafc"] * len(metrics)]
                        + [["#ffffff"] * len(metrics)] * len(df),
                        font=dict(size=12, color="#1e293b"),
                        align="left",
                        height=28,
                    ),
                )
            ]
        )
        fig.update_layout(
            template=_PLOTLY_TEMPLATE,
            title="Scorecard :: KPI per Progression Script",
            height=max(320, 60 + 30 * len(metrics)),
            margin=dict(l=20, r=20, t=70, b=20),
        )
        return fig

    # ── §2 age curve ──────────────────────────────────────────────────
    def cmp_age_curve(self) -> go.Figure:
        """Overlaid mean-Δ age curves with 95% CI bands + peak markers.

        TUNING TAKEAWAY: THE comparison. Curves crossing zero further right
        peak later; flatter post-peak tails decline gentler.
        """
        fig = go.Figure()
        for r in self._runs:
            pa = r.per_age
            if pa.empty:
                continue
            c = self._colors[r.label]
            fig.add_trace(
                go.Scatter(
                    x=pd.concat([pa["Age"], pa["Age"][::-1]]),
                    y=pd.concat([pa["CI_hi"], pa["CI_lo"][::-1]]),
                    fill="toself",
                    fillcolor=hex_to_rgba(c, 0.12),
                    line=dict(color="rgba(0,0,0,0)"),
                    showlegend=False,
                    hoverinfo="skip",
                )
            )
            fig.add_trace(
                go.Scatter(
                    x=pa["Age"],
                    y=pa["Mean"],
                    mode="lines+markers",
                    name=r.label,
                    line=dict(color=c, width=3),
                    marker=dict(size=6, color=c),
                    hovertemplate=(
                        f"<b>{r.label}</b><br>Age %{{x}}<br>"
                        f"Mean Δ %{{y:+.2f}}<extra></extra>"
                    ),
                )
            )
            if not np.isnan(r.peak_age):
                fig.add_vline(
                    x=r.peak_age,
                    line_dash="dot",
                    line_color=c,
                    line_width=1.5,
                    opacity=0.6,
                )
        fig.add_hline(y=0, line_width=1.5, line_color="#1e293b")
        fig.update_layout(
            template=_PLOTLY_TEMPLATE,
            title="Age Curve Overlay :: Mean OVR Δ per Age (±95% CI)",
            xaxis_title="Age",
            yaxis_title="Mean OVR Δ",
            hovermode="x unified",
        )
        return fig

    def cmp_sigma_by_age(self) -> go.Figure:
        """Per-age σ(Δ) overlay: where each script injects its variance.

        TUNING TAKEAWAY: a script with a σ spike in the young tier is
        loading its RNG (or god-progs) there; flat σ = uniform noise.
        """
        fig = go.Figure()
        for r in self._runs:
            pa = r.per_age
            if pa.empty:
                continue
            c = self._colors[r.label]
            fig.add_trace(
                go.Scatter(
                    x=pa["Age"],
                    y=pa["Std"],
                    mode="lines+markers",
                    name=r.label,
                    line=dict(color=c, width=2.5),
                    marker=dict(size=5, color=c),
                    hovertemplate=(
                        f"<b>{r.label}</b><br>Age %{{x}}<br>"
                        f"σ(Δ) %{{y:.2f}}<extra></extra>"
                    ),
                )
            )
        fig.update_layout(
            template=_PLOTLY_TEMPLATE,
            title="Run-to-Run σ(Δ) by Age",
            xaxis_title="Age",
            yaxis_title="σ(Δ)",
            hovermode="x unified",
        )
        return fig

    # ── §3 separation ─────────────────────────────────────────────────
    def cmp_separation_bands(self) -> go.Figure:
        """Grouped bars: production-tercile Δ gap per age band per script.

        TUNING TAKEAWAY: the fairness fingerprint. Tall positive prime-age
        bars = studs out-progress scrubs; bars fading to ~0 with age = the
        talent effect correctly stops lifting old players.
        """
        bands = [f"{a}-{b}" for a, b in Config.SEP_BANDS]
        fig = go.Figure()
        for r in self._runs:
            c = self._colors[r.label]
            fig.add_trace(
                go.Bar(
                    name=r.label,
                    x=bands,
                    y=[r.sep_bands.get(b, np.nan) for b in bands],
                    marker_color=c,
                    opacity=0.9,
                    hovertemplate=(
                        f"<b>{r.label}</b><br>Ages %{{x}}<br>"
                        f"top−bottom tercile Δ gap: %{{y:+.3f}}<extra></extra>"
                    ),
                )
            )
        fig.add_hline(y=0, line_width=1.5, line_color="#1e293b")
        prod = next((r.prod_label for r in self._runs if r.prod_label), "production")
        fig.update_layout(
            template=_PLOTLY_TEMPLATE,
            title=f"Talent Separation by Age Band ({prod} terciles)",
            xaxis_title="Age band",
            yaxis_title="Top − Bottom tercile mean Δ",
            barmode="group",
        )
        return fig

    def cmp_delta_vs_baseline(self) -> go.Figure:
        """Mean Δ vs baseline-OVR bin, overlaid. Shows the ceiling bite.

        TUNING TAKEAWAY: where each script's curve turns down as baseline
        rises is where its cap starts pulling. A script flat into high
        baselines isn't capping; one that dives is capping hard.
        """
        edges = np.arange(40, 95, 5.0)
        centers = (edges[:-1] + edges[1:]) / 2
        fig = go.Figure()
        for r in self._runs:
            ppl = r.per_player.dropna(subset=["Baseline", "MeanDelta"])
            if ppl.empty:
                continue
            c = self._colors[r.label]
            binidx = np.digitize(ppl["Baseline"].values, edges) - 1
            ys = []
            for i in range(len(centers)):
                m = binidx == i
                ys.append(ppl["MeanDelta"].values[m].mean() if m.sum() >= 3 else np.nan)
            fig.add_trace(
                go.Scatter(
                    x=centers,
                    y=ys,
                    mode="lines+markers",
                    name=r.label,
                    line=dict(color=c, width=2.5),
                    marker=dict(size=6, color=c),
                    connectgaps=False,
                    hovertemplate=(
                        f"<b>{r.label}</b><br>baseline ~%{{x:.0f}}<br>"
                        f"mean Δ %{{y:+.2f}}<extra></extra>"
                    ),
                )
            )
        fig.add_hline(y=0, line_width=1.5, line_color="#1e293b")
        fig.update_layout(
            template=_PLOTLY_TEMPLATE,
            title="Mean Δ vs Baseline OVR (fixed 5-pt bins)",
            xaxis_title="Baseline OVR (bin center)",
            yaxis_title="Mean OVR Δ",
            hovermode="x unified",
        )
        return fig

    # ── §4 ceiling / outcomes ─────────────────────────────────────────
    def cmp_improve_probability(self) -> go.Figure:
        """P(Δ>0) vs baseline OVR, one smoothed line per script.

        TUNING TAKEAWAY: the height and drop-off point of each curve says
        how generous the script is and where its ceiling starts to bite.
        """
        edges = np.arange(40, 95, 5.0)
        centers = (edges[:-1] + edges[1:]) / 2
        fig = go.Figure()
        for r in self._runs:
            ppl = r.per_player.dropna(subset=["Baseline", "PctPositive"])
            if ppl.empty:
                continue
            c = self._colors[r.label]
            binidx = np.digitize(ppl["Baseline"].values, edges) - 1
            ys = []
            for i in range(len(centers)):
                m = binidx == i
                ys.append(
                    ppl["PctPositive"].values[m].mean() if m.sum() >= 3 else np.nan
                )
            fig.add_trace(
                go.Scatter(
                    x=centers,
                    y=ys,
                    mode="lines+markers",
                    name=r.label,
                    line=dict(color=c, width=2.5),
                    marker=dict(size=6, color=c),
                    connectgaps=False,
                    hovertemplate=(
                        f"<b>{r.label}</b><br>baseline ~%{{x:.0f}}<br>"
                        f"P(Δ>0) %{{y:.0%}}<extra></extra>"
                    ),
                )
            )
        fig.add_hline(y=0.5, line_width=1, line_dash="dot", line_color="#94a3b8")
        fig.update_layout(
            template=_PLOTLY_TEMPLATE,
            title="Probability of Improvement vs Baseline OVR",
            xaxis_title="Baseline OVR (bin center)",
            yaxis_title="P(Δ > 0)",
            yaxis_tickformat=".0%",
            hovermode="x unified",
        )
        return fig

    # ── §5 noise ──────────────────────────────────────────────────────
    def cmp_noise_kde(self) -> go.Figure:
        """KDE of per-player StdDelta, overlaid.

        TUNING TAKEAWAY: the run-to-run noise distribution. A script shifted
        right is noisier per player; a long right tail means a subset of
        players are being thrown around by RNG.
        """
        fig = go.Figure()
        all_sig = (
            np.concatenate(
                [
                    r.per_player["StdDelta"].dropna().values
                    for r in self._runs
                    if not r.per_player.empty
                ]
            )
            if self._runs
            else np.array([])
        )
        if all_sig.size == 0:
            return _empty_fig("Noise KDE")
        xgrid = np.linspace(0, float(np.nanpercentile(all_sig, 99)) + 0.5, 300)
        for r in self._runs:
            vals = r.per_player["StdDelta"].dropna().values
            if len(vals) < Config.MIN_KDE_SAMPLES or np.std(vals) < 1e-9:
                continue
            c = self._colors[r.label]
            kde = gaussian_kde(vals, bw_method="scott")(xgrid)
            fig.add_trace(
                go.Scatter(
                    x=xgrid,
                    y=kde,
                    mode="lines",
                    name=r.label,
                    fill="tozeroy",
                    fillcolor=hex_to_rgba(c, 0.12),
                    line=dict(color=c, width=2.5, shape="spline"),
                    hovertemplate=(
                        f"<b>{r.label}</b><br>σΔ %{{x:.2f}}<br>"
                        f"density %{{y:.3f}}<extra></extra>"
                    ),
                )
            )
            fig.add_vline(
                x=float(np.median(vals)),
                line_dash="dash",
                line_color=c,
                line_width=1.5,
                opacity=0.6,
            )
        fig.update_layout(
            template=_PLOTLY_TEMPLATE,
            title="Per-Player Run-to-Run Noise σ(Δ) (dashed = median)",
            xaxis_title="σ(Δ) per player",
            yaxis_title="Density",
            hovermode="x unified",
        )
        return fig

    def cmp_icc_bars(self) -> go.Figure:
        """Grouped ICC bars: OVR + median-attribute per script.

        TUNING TAKEAWAY: signal share. Higher = more deterministic (player
        identity dominates); lower = RNG dominates. Compare against your
        target band (roughly 0.4-0.85).
        """
        fig = go.Figure()
        cats = ["ICC(Ovr)", "ICC(median attr)"]
        for r in self._runs:
            c = self._colors[r.label]
            fig.add_trace(
                go.Bar(
                    name=r.label,
                    x=cats,
                    y=[r.icc_ovr, r.icc_attr_median],
                    marker_color=c,
                    opacity=0.9,
                    hovertemplate=(
                        f"<b>{r.label}</b><br>%{{x}}: %{{y:.3f}}<extra></extra>"
                    ),
                )
            )
        for yv, lbl in [(0.4, "low"), (0.7, "high")]:
            fig.add_hline(
                y=yv,
                line_dash="dot",
                line_color="#94a3b8",
                annotation_text=lbl,
                annotation_position="right",
            )
        fig.update_layout(
            template=_PLOTLY_TEMPLATE,
            title="ICC :: Signal Share (OVR & median attribute)",
            xaxis_title="",
            yaxis_title="ICC",
            barmode="group",
            yaxis_range=[0, 1],
        )
        return fig

    # ── §6 fingerprint ────────────────────────────────────────────────
    def cmp_attr_fingerprint(self) -> go.Figure:
        """Per-attribute OVR-weighted mean movement, grouped by script.

        TUNING TAKEAWAY: two scripts can share an age curve but move
        different attributes to get there. Bars are coef·meanΔ (OVR-impact
        units), so height is directly comparable across attributes.
        """
        coef_map = dict(zip(Config.OVR_CALC_ORDER, Config.OVR_COEFFS))
        # Union of varying attrs across runs, ordered by group then coef
        attrs = sorted(
            {a for r in self._runs for a in r.varying_attrs},
            key=lambda a: (
                ["Physical", "Shooting", "Mental", "Skill", "Fixed", "Other"].index(
                    _attr_group(a)
                ),
                -coef_map.get(a, 0.0),
            ),
        )
        if not attrs:
            return _empty_fig("Attribute Fingerprint")
        fig = go.Figure()
        for r in self._runs:
            if r.attr_deltas.empty:
                continue
            c = self._colors[r.label]
            ys = []
            for a in attrs:
                if a in r.attr_deltas.columns:
                    ys.append(float(r.attr_deltas[a].mean()) * coef_map.get(a, 0.0))
                else:
                    ys.append(np.nan)
            fig.add_trace(
                go.Bar(
                    name=r.label,
                    x=attrs,
                    y=ys,
                    marker_color=c,
                    opacity=0.9,
                    hovertemplate=(
                        f"<b>{r.label}</b><br>%{{x}}<br>"
                        f"OVR-weighted Δ %{{y:+.3f}}<extra></extra>"
                    ),
                )
            )
        fig.add_hline(y=0, line_width=1.5, line_color="#1e293b")
        fig.update_layout(
            template=_PLOTLY_TEMPLATE,
            title="Attribute Fingerprint :: coef × mean Δ (OVR-impact units)",
            xaxis_title="Attribute (group-ordered)",
            yaxis_title="OVR points contributed",
            barmode="group",
        )
        return fig


class ComparisonDashboard:
    """Comparison dashboard (thin wrapper over the shared lazy renderer)."""

    def __init__(self, charts: Dict[str, go.Figure], runs: List[RunSummary]):
        self._charts = charts
        self._runs = runs

    def render(self, path: Path) -> None:
        n_scripts = len(self._runs)
        peaks = [r.peak_age for r in self._runs if not np.isnan(r.peak_age)]
        peak_span = f"{min(peaks):.1f}-{max(peaks):.1f}" if peaks else "-"
        labels = ", ".join(r.label for r in self._runs)

        _render_html(
            path,
            title="Progbox · Progression Script Comparison",
            subtitle=f"{n_scripts} scripts: {labels}",
            hero_h1="Progression Script Comparison",
            hero_p=(
                "Head-to-head across progression scripts. The scorecard "
                "distills every tuning KPI to one number per script; the "
                "overlays show the shape behind each number."
            ),
            nav=[(sec["id"], sec["title"]) for sec in CMP_SECTIONS],
            stat_cards=[
                ("Scripts", f"{n_scripts}", None),
                ("Peak-age span", peak_span, None),
                ("Players", f"{self._runs[0].n_players:,}", None),
                ("Runs each", f"{self._runs[0].n_runs:,}", None),
            ],
            sections=CMP_SECTIONS,
            charts=self._charts,
        )


# PIPELINE
def generate_analysis(
    run_dir: Optional[str] = None, full_excel: bool = False, no_excel: bool = False
) -> None:
    """Single-run pipeline. Discovers a run, processes it, writes outputs."""
    loader = DataLoader()
    base = Path(run_dir) if run_dir else loader.find_latest_run()
    raw_dir = base / "raw"

    logger.info("=" * 70)
    logger.info("Progbox Tuning Analysis Pipeline · single-run")
    logger.info("=" * 70)

    sim = loader.load_sim(raw_dir)
    baseline = loader.load_baseline(raw_dir)
    meta = loader.load_metadata(base)
    label = (
        (meta.get("progression", {}) or {}).get("name")
        or (meta.get("progression", {}) or {}).get("id")
        or ""
    )
    logger.info(
        f"Detected {sim['Run'].nunique()} runs x "
        f"{sim['PlayerID'].nunique()} players = {len(sim):,} rows"
    )

    ds = Dataset(sim, baseline)

    if not no_excel:
        ExcelWorkbook(ds, full_excel=full_excel).render(base / "analysis.xlsx")

    logger.info("Building charts...")
    builder = ChartBuilder(ds)
    charts = builder.build_all()

    # Inject the interactive per-player explorer widget (rendered as raw HTML).
    logger.info("  [widget] player_explorer")
    charts["player-explorer/player_explorer"] = builder.build_player_explorer()

    _validate_chart_registry()

    HTMLDashboard(charts, ds, label=label).render(base / "analysis_dashboard.html")

    logger.info("=" * 70)
    logger.info(
        f"Done · charts: {len(charts)} · players: {sim['PlayerID'].nunique()} · "
        f"runs: {sim['Run'].nunique()}"
    )
    logger.info("=" * 70)


def generate_comparison(
    run_dirs: List[str], ceiling: float = Config.CEILING_OVR
) -> None:
    """Comparison pipeline. Loads N runs, emits comparison dashboard + CSV."""
    logger.info("=" * 70)
    logger.info(f"Progbox Comparison Pipeline · {len(run_dirs)} scripts")
    logger.info("=" * 70)

    runs: List[RunSummary] = []
    for rd in run_dirs:
        p = Path(rd)
        logger.info(f"Summarizing {p} ...")
        runs.append(RunSummary(p, ceiling=ceiling))

    builder = ComparisonBuilder(runs)
    logger.info("Building comparison charts...")
    charts = builder.build_all()

    # Output next to the first run dir
    out_base = Path(run_dirs[0])
    scorecard = builder.scorecard_df()
    csv_path = out_base / "comparison_scorecard.csv"
    scorecard.to_csv(csv_path, index=False)
    logger.info(f"Scorecard CSV --> {csv_path}")

    html_path = out_base / "comparison_dashboard.html"
    ComparisonDashboard(charts, runs).render(html_path)

    logger.info("=" * 70)
    logger.info("Scorecard:\n" + scorecard.to_string(index=False))
    logger.info("=" * 70)


def main() -> None:
    p = argparse.ArgumentParser(
        description="Progbox Monte Carlo analysis. One run --> seven-section "
        "dashboard; multiple runs --> head-to-head comparison."
    )
    p.add_argument(
        "run_dirs",
        nargs="*",
        default=None,
        help="Run directories. 0 --> auto-discover latest (single mode); "
        "1 --> single-run dashboard; 2+ --> comparison dashboard.",
    )
    p.add_argument(
        "--ceiling",
        type=float,
        default=Config.CEILING_OVR,
        help=f"OVR ceiling for the over-cap check "
        f"(default {Config.CEILING_OVR:g}, = softCeil+band).",
    )
    p.add_argument(
        "--full-excel",
        action="store_true",
        help="Include the heavy 'All Runs' sheet (duplicates "
        "outputs.csv; adds ~1-2 min at large scale).",
    )
    p.add_argument(
        "--no-excel",
        action="store_true",
        help="Skip the Excel workbook entirely (HTML only).",
    )
    args = p.parse_args()

    try:
        dirs = args.run_dirs or []
        if len(dirs) >= 2:
            generate_comparison(dirs, ceiling=args.ceiling)
        else:
            generate_analysis(
                dirs[0] if dirs else None,
                full_excel=args.full_excel,
                no_excel=args.no_excel,
            )
    except Exception as e:
        logger.error(f"Fatal error: {e}", exc_info=True)
        sys.exit(1)


if __name__ == "__main__":
    main()
